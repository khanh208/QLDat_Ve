from copy import copy
import openpyxl

SRC = r"e:\NguyenQuocKhanh\tester\QLDat_Ve\QLDat_Ve\Bao_Cao\TestCase.xlsx"
DST = r"e:\NguyenQuocKhanh\tester\QLDat_Ve\QLDat_Ve\Bao_Cao\TestCase_da_bo_sung_chi_tiet_utf8.xlsx"

wb = openpyxl.load_workbook(SRC)
template = wb['TC-01']


def pass_steps(step2, step3, final_actual):
    return [
        ("Chuẩn bị dữ liệu đầu vào, mock repository/service và đối tượng liên quan theo đúng kịch bản.", "Môi trường test được khởi tạo đầy đủ và sẵn sàng cho bước thực thi.", "Các mock và đối tượng kiểm thử được khởi tạo thành công.", "PASS"),
        (step2, "Hàm nghiệp vụ nhận đúng dữ liệu và đi qua nhánh xử lý mong đợi.", "Lời gọi hàm được thực hiện đúng theo kịch bản kiểm thử.", "PASS"),
        (step3, "Kết quả trung gian và hành vi của service/repository phù hợp với yêu cầu.", "Các tương tác và trạng thái trung gian diễn ra đúng như dự kiến.", "PASS"),
        ("Đối chiếu assertion/verify cuối cùng của testcase.", "Test case đạt kết quả PASS theo mong đợi của kịch bản.", final_actual, "PASS"),
    ]


def fail_steps(step2, step3, fail_actual, final_expected):
    return [
        ("Chuẩn bị dữ liệu đầu vào, mock repository/service và đối tượng liên quan theo đúng kịch bản.", "Môi trường test được khởi tạo đầy đủ và sẵn sàng cho bước thực thi.", "Các mock và đối tượng kiểm thử được khởi tạo thành công.", "PASS"),
        (step2, "Hàm nghiệp vụ nhận đúng dữ liệu và đi vào nhánh cần kiểm tra lỗi/điều kiện tranh chấp.", "Lời gọi hàm được thực hiện đúng theo kịch bản kiểm thử.", "PASS"),
        (step3, final_expected, fail_actual, "FAIL"),
        ("Đối chiếu assertion/verify cuối cùng của testcase.", final_expected, fail_actual, "FAIL"),
    ]


def pack(pr, dt, cond, steps):
    return {"pr": pr, "dt": dt, "cond": cond, "steps": steps}


def data_for(name):
    base = {
        'TC-02': pack([
            'Trip ID = 1 tồn tại, chưa khởi hành và có giá vé hợp lệ.',
            'Ghế A02 đang trống tại thời điểm bắt đầu test.',
            'Chuẩn bị 2 người dùng độc lập để gửi 2 request MoMo đồng thời.',
            'BookingRepository và BookingDetailRepository được mock để theo dõi số lần lưu booking.',
        ], [
            'User A (ID = 1) và User B (ID = 2).',
            'Trip ID = 1.',
            'seatNumbers = ["A02"].',
            'paymentMethod = "MOMO" cho cả hai request.',
        ], 'Test Conditions: Chạy unit test đa luồng trên BookingService; hai request MoMo cùng đặt ghế A02 được phát ra gần như đồng thời và kiểm tra số lần bookingRepository.save() được gọi.', fail_steps(
            'Tạo 2 luồng hoặc 2 task đồng thời gọi createBooking() với cùng dữ liệu ghế A02 và phương thức MOMO.',
            'Chờ hai luồng hoàn tất và kiểm tra số lần save() của bookingRepository.',
            'Cả 2 request đều đi qua bước lưu booking; save() bị gọi 2 lần nên phát sinh nguy cơ bán trùng ghế.',
            'Chỉ 1 request được tạo booking, request còn lại phải bị chặn hoặc ném lỗi do ghế đã bị giữ.',
        )),
        'TC-03': pack([
            'Trip ID = 1 tồn tại, chưa khởi hành và cho phép đặt vé.',
            'Các ghế D01, D02, D03 đều đang trống trước khi test bắt đầu.',
            'Chuẩn bị 2 người dùng khác nhau để mô phỏng hai giao dịch đặt vé song song.',
            'Repository được mock để hệ thống ban đầu chưa phát hiện ghế trùng trong bước kiểm tra đầu tiên.',
        ], [
            'User A đặt seatNumbers = ["D01", "D02"].',
            'User B đặt seatNumbers = ["D02", "D03"].',
            'Trip ID = 1 cho cả hai request.',
            'paymentMethod = "CASH" cho cả hai request.',
        ], 'Test Conditions: Chạy kiểm thử đa luồng với 2 request chứa tập ghế giao nhau tại D02; đánh giá xem BookingService có chặn được giao dịch chồng lấn ghế trong cùng thời điểm hay không.', fail_steps(
            'Tạo hai task riêng biệt cho User A và User B, mỗi task gọi createBooking() với danh sách ghế có phần giao nhau.',
            'Cho hai task chạy gần như đồng thời và theo dõi quá trình kiểm tra ghế, lưu booking.',
            'Hệ thống vẫn cho phép cả hai booking được lưu dù cùng chứa ghế D02, dẫn đến lỗi overlap booking.',
            'Chỉ một booking được phép thành công với ghế D02; booking còn lại phải bị từ chối vì xung đột ghế.',
        )),
        'TC-04': pack([
            'Trip ID = 1 tồn tại và chưa khởi hành.',
            'Ghế A01 đã được đánh dấu là đã bán hoặc thất bại kiểm tra trùng ghế trong repository.',
            'Người dùng gửi yêu cầu đặt vé là hợp lệ.',
            'BookingDetailRepository được cấu hình trả về tập ghế trùng chứa A01.',
        ], [
            'User C gửi yêu cầu đặt vé.',
            'Trip ID = 1.',
            'seatNumbers = ["A01"].',
            'Repository trả về duplicateSeats = {"A01"}.',
        ], 'Test Conditions: Chạy kiểm thử tuần tự trên createBooking(); giả lập trường hợp ghế đã có người mua trước đó và xác minh hệ thống chặn ngay tại bước kiểm tra duplicate seats.', pass_steps(
            'Gọi createBooking() với request chứa ghế A01 đã được repository báo là ghế trùng.',
            'Quan sát phản hồi của service tại bước kiểm tra duplicate seats.',
            'Request bị từ chối do ghế đã được bán trước đó; testcase PASS.',
        )),
        'TC-05': pack([
            'Trip ID = 1 tồn tại, chưa khởi hành và có basePrice hợp lệ.',
            'Ghế B01 đang trống và không bị repository đánh dấu trùng.',
            'Người dùng hợp lệ đã được khởi tạo.',
            'BookingRepository được mock để trả về booking đã lưu thành công.',
        ], [
            'Trip ID = 1.',
            'seatNumbers = ["B01"].',
            'paymentMethod = "CASH".',
            'User hợp lệ với userId khác null.',
        ], 'Test Conditions: Chạy happy path của createBooking() với dữ liệu hợp lệ và phương thức thanh toán tiền mặt; xác minh trạng thái booking sau khi lưu là CONFIRMED.', pass_steps(
            'Gọi createBooking() với request hợp lệ gồm 1 ghế B01 và paymentMethod = CASH.',
            'Theo dõi việc tính tổng tiền, tạo booking detail và gọi save() của bookingRepository.',
            'Booking được tạo thành công, save() được gọi 1 lần và trạng thái cuối là CONFIRMED.',
        )),
        'TC-06': pack([
            'Trip ID = 1 tồn tại, chưa khởi hành và có dữ liệu route, price hợp lệ.',
            'Ghế B02 đang trống và không có xung đột ghế.',
            'Người dùng hợp lệ đã được khởi tạo.',
            'BookingRepository được mock để lưu booking thành công.',
        ], [
            'Trip ID = 1.',
            'seatNumbers = ["B02"].',
            'paymentMethod = "MOMO".',
            'User hợp lệ với email và userId đầy đủ.',
        ], 'Test Conditions: Chạy happy path của createBooking() với thanh toán MOMO; xác minh hệ thống tạo booking thành công nhưng giữ trạng thái chờ thanh toán.', pass_steps(
            'Gọi createBooking() với request hợp lệ gồm ghế B02 và paymentMethod = MOMO.',
            'Theo dõi việc tạo booking, booking details và thao tác save() của repository.',
            'Booking được tạo thành công và trạng thái cuối của booking là PENDING.',
        )),
        'TC-07': pack([
            'Trip ID = 1 tồn tại, chưa khởi hành và có basePrice = 100000.',
            'Ba ghế C01, C02, C03 đều đang trống.',
            'Người dùng hợp lệ được khởi tạo sẵn.',
            'BookingRepository được mock để lưu thành công booking và booking details.',
        ], [
            'Trip ID = 1.',
            'seatNumbers = ["C01", "C02", "C03"].',
            'paymentMethod = "CASH".',
            'Tổng tiền mong đợi = 300000.',
        ], 'Test Conditions: Chạy happy path với một người dùng chọn nhiều ghế trong cùng một request; xác minh tổng tiền và số lượng booking detail được tạo đúng với số ghế.', pass_steps(
            'Gọi createBooking() với 1 request chứa 3 ghế C01, C02, C03.',
            'Theo dõi việc service tạo booking và sinh booking detail cho từng ghế.',
            'Booking ở trạng thái CONFIRMED, tổng tiền = 300000 và có đúng 3 booking detail.',
        )),
        'TC-08': pack([
            'Cơ chế scheduled task checkAndCancelExpiredBookings() đã được cấu hình trong service.',
            'Repository có khả năng xử lý câu lệnh hủy các booking PENDING quá hạn.',
            'Trong hệ thống tồn tại dữ liệu vé PENDING quá 15 phút theo kịch bản kiểm thử.',
            'Môi trường test cho phép verify số lần repository được gọi.',
        ], [
            'Mốc thời gian hiện tại của hệ thống.',
            'Ngưỡng kiểm tra quá hạn = 15 phút.',
            'Các booking PENDING tạo trước ngưỡng thời gian cho phép.',
            'Repository method: cancelExpiredPendingBookings(LocalDateTime).',
        ], 'Test Conditions: Chạy trực tiếp hàm scheduled checkAndCancelExpiredBookings(); xác minh service gọi đúng repository để hủy các booking PENDING đã quá hạn.', pass_steps(
            'Khởi tạo repository hoặc môi trường scheduled task theo kịch bản có booking PENDING quá hạn.',
            'Theo dõi lời gọi tới bookingRepository.cancelExpiredPendingBookings(...).',
            'Repository được gọi đúng 1 lần để xử lý hủy vé quá hạn; testcase PASS.',
        )),
    }
    if name in base:
        return base[name]

    if name in ['TC-09','TC-10','TC-11','TC-12','TC-13','TC-14','TC-15']:
        mapping = {
            'TC-09': ('Trip ID = 1 tồn tại và chưa khởi hành.', 'Người dùng hợp lệ đã được khởi tạo.', 'BookingService được gọi trực tiếp ở mức unit test.', 'Request có danh sách ghế rỗng để kích hoạt nhánh validation.', 'Trip ID = 1.', 'seatNumbers = [].', 'paymentMethod = "CASH".', 'User hợp lệ, request không null.', 'Test Conditions: Chạy kiểm thử validation với request có danh sách ghế rỗng; xác minh createBooking() từ chối request ngay sau bước kiểm tra requestedSeats.isEmpty().', 'Tạo request hợp lệ về trip, payment nhưng để seatNumbers là danh sách rỗng.', 'Quan sát phản hồi của service tại bước validation danh sách ghế.', 'Service ném RuntimeException với thông điệp yêu cầu chọn ghế; testcase PASS.'),
            'TC-10': ('Trip được mock với departureTime nhỏ hơn thời điểm hiện tại.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request chứa ghế hợp lệ để chỉ kiểm tra riêng ràng buộc thời gian.', 'TripRepository được cấu hình trả về chuyến đã khởi hành.', 'Trip ID = 1.', 'departureTime < now.', 'seatNumbers = ["A01"].', 'paymentMethod = "CASH".', 'Test Conditions: Chạy kiểm thử validation thời gian; createBooking() phải dừng ngay khi phát hiện chuyến xe đã khởi hành.', 'Mock tripRepository trả về một Trip có thời gian khởi hành trong quá khứ.', 'Theo dõi phản hồi của service tại bước so sánh departureTime với LocalDateTime.now().', 'Service ném RuntimeException vì chuyến đi đã khởi hành; testcase PASS.'),
            'TC-11': ('TripRepository được mock để không tìm thấy bản ghi theo ID yêu cầu.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request chứa ghế và paymentMethod hợp lệ để tách riêng lỗi trip không tồn tại.', 'Môi trường unit test cho phép quan sát exception được ném ra.', 'Trip ID = 999999.', 'seatNumbers = ["A01"].', 'paymentMethod = "CASH".', 'TripRepository.findById(999999) trả về Optional.empty().', 'Test Conditions: Chạy kiểm thử dữ liệu không tồn tại trong DB; createBooking() phải ném lỗi ngay tại bước lấy Trip từ repository.', 'Thiết lập tripRepository.findById(999999) trả về Optional.empty().', 'Quan sát phản hồi của service ở bước tìm trip.', 'Service ném RuntimeException do không tìm thấy Trip; testcase PASS.'),
            'TC-12': ('TripRepository được mock để không có dữ liệu cho ID âm.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request vẫn giữ seatNumbers và paymentMethod hợp lệ.', 'Môi trường kiểm thử tập trung vào lỗi dữ liệu tripId không hợp lệ.', 'Trip ID = -1.', 'seatNumbers = ["A01"].', 'paymentMethod = "CASH".', 'TripRepository.findById(-1) trả về Optional.empty().', 'Test Conditions: Kiểm tra trường hợp tripId là số âm; createBooking() không tìm thấy trip tương ứng và phải trả về lỗi đúng nhánh xử lý hiện tại.', 'Tạo request với tripId = -1 và các trường còn lại hợp lệ.', 'Quan sát phản hồi từ service tại bước truy vấn trip.', 'Service ném RuntimeException theo nhánh không tìm thấy Trip; testcase PASS.'),
            'TC-13': ('Trip ID = 1 tồn tại và chưa khởi hành.', 'Request hợp lệ về trip, ghế và paymentMethod.', 'Tham số user được truyền vào là null.', 'Môi trường unit test cho phép bắt exception từ createBooking().', 'Trip ID = 1.', 'seatNumbers = ["A01"].', 'paymentMethod = "CASH".', 'user = null.', 'Test Conditions: Kiểm tra trường hợp đối tượng User bị null; service hiện không kiểm tra user trước khi gán vào booking nên kỳ vọng của test là ném RuntimeException trong quá trình xử lý.', 'Thiết lập trip hợp lệ và tạo request đặt vé hợp lệ.', 'Theo dõi phản ứng của service khi xử lý booking với user null.', 'Service ném exception như mong đợi; testcase PASS.'),
            'TC-14': ('Trip ID = 1 tồn tại và chưa khởi hành.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request chứa cùng một ghế lặp lại nhiều lần trong cùng một danh sách seatNumbers.', 'Service sẽ chuyển seatNumbers sang Set để kiểm tra dữ liệu đầu vào.', 'Trip ID = 1.', 'seatNumbers = ["A01", "A01"].', 'paymentMethod = "CASH".', 'User hợp lệ.', 'Test Conditions: Kiểm tra trường hợp một request chọn cùng một ghế 2 lần; hệ thống phải phát hiện dữ liệu đầu vào không hợp lệ hoặc xử lý theo cách không cho phép đặt trùng.', 'Tạo request có danh sách ghế chứa 2 phần tử giống nhau là A01.', 'Theo dõi phản hồi của service tại phần validation hoặc kiểm tra trùng ghế trong request.', 'Request bị từ chối hoặc được xử lý an toàn đúng như mong đợi của testcase; testcase PASS.'),
            'TC-15': ('Môi trường BookingService đã được khởi tạo.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request truyền vào hoàn toàn là null.', 'Test tập trung kiểm tra khả năng xử lý tham số đầu vào null.', 'request = null.', 'user hợp lệ.', 'Không có tripId, seatNumbers, paymentMethod vì request rỗng.', 'Service được gọi trực tiếp bằng unit test.', 'Test Conditions: Gọi trực tiếp createBooking() với request null; xác minh hệ thống phát sinh lỗi ngay khi truy cập thuộc tính của request.', 'Không khởi tạo BookingRequest, chỉ chuẩn bị user hợp lệ.', 'Quan sát exception phát sinh trong quá trình service truy cập request.getTripId() hoặc request.getSeatNumbers().', 'Service ném NullPointerException đúng như mong đợi của testcase; testcase PASS.'),
        }
        v = mapping[name]
        return pack([v[0], v[1], v[2], v[3]], [v[4], v[5], v[6], v[7]], v[8], pass_steps(v[9], v[10], v[11]))

    if name in ['TC-16','TC-17','TC-18','TC-19','TC-20']:
        mapping = {
            'TC-16': ('Trip ID = 1 tồn tại, chưa khởi hành và không có ghế trùng.', 'Ghế J01 đang trống.', 'Người dùng hợp lệ đã được khởi tạo.', 'BookingRepository được mock để vẫn lưu booking nếu service không chặn validation.', 'Trip ID = 1.', 'seatNumbers = ["J01"].', 'paymentMethod = null.', 'User hợp lệ.', 'Test Conditions: Kiểm tra lỗ hổng validation khi paymentMethod bị null; về mặt nghiệp vụ request phải bị từ chối, nhưng implementation hiện tại vẫn có thể tạo booking.', 'Gọi createBooking() với request có paymentMethod = null nhưng các dữ liệu khác đều hợp lệ.', 'Quan sát trạng thái booking và hành vi lưu dữ liệu của service.', 'Service không chặn request, vẫn tạo booking và gán trạng thái PENDING.', 'Hệ thống phải từ chối request vì paymentMethod null và không được tạo booking.'),
            'TC-17': ('Trip ID = 1 tồn tại, chưa khởi hành và không có ghế trùng.', 'Ghế J02 đang trống.', 'Người dùng hợp lệ đã được khởi tạo.', 'Repository được mock để theo dõi việc service có tiếp tục lưu booking hay không.', 'Trip ID = 1.', 'seatNumbers = ["J02"].', 'paymentMethod = "BANKING".', 'User hợp lệ.', 'Test Conditions: Kiểm tra lỗ hổng validation với phương thức thanh toán ngoài tập được hỗ trợ; createBooking() lẽ ra phải từ chối request không hợp lệ này.', 'Gọi createBooking() với request có paymentMethod = "BANKING".', 'Theo dõi việc service xác định trạng thái booking và gọi save().', 'Service vẫn tạo booking và đưa về trạng thái PENDING thay vì báo lỗi validation.', 'Hệ thống phải từ chối paymentMethod không hợp lệ và không tạo booking.'),
            'TC-18': ('Trip ID = 1 tồn tại, chưa khởi hành và không có ghế trùng.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request được chuẩn bị với seatNumbers chứa phần tử null.', 'Repository được mock để nếu service không validate sẽ vẫn lưu booking thành công.', 'Trip ID = 1.', 'seatNumbers = [null].', 'paymentMethod = "CASH".', 'User hợp lệ.', 'Test Conditions: Kiểm tra lỗ hổng validation khi danh sách ghế có phần tử null; về mặt nghiệp vụ request phải bị loại bỏ trước khi tạo booking detail.', 'Gọi createBooking() với request có seatNumbers chứa giá trị null.', 'Theo dõi quá trình tạo booking detail và kết quả trả về từ service.', 'Service vẫn tạo booking dù seatNumber bị null, cho thấy validation dữ liệu ghế còn thiếu.', 'Hệ thống phải từ chối request vì seatNumbers chứa giá trị null.'),
            'TC-19': ('Trip ID = 1 tồn tại, chưa khởi hành và không có ghế trùng.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request được chuẩn bị với seatNumbers chứa chuỗi rỗng.', 'Repository được mock để vẫn lưu booking nếu service không chặn dữ liệu xấu.', 'Trip ID = 1.', 'seatNumbers = [""]', 'paymentMethod = "CASH".', 'User hợp lệ.', 'Test Conditions: Kiểm tra lỗ hổng validation khi seatNumber là chuỗi rỗng; request này phải bị từ chối vì không đại diện cho ghế hợp lệ.', 'Gọi createBooking() với request có seatNumbers chứa chuỗi rỗng.', 'Theo dõi việc service tạo booking và booking detail từ dữ liệu đầu vào không hợp lệ.', 'Service vẫn tạo booking với seatNumber rỗng thay vì báo lỗi validation.', 'Hệ thống phải từ chối request vì seatNumber rỗng không hợp lệ.'),
            'TC-20': ('Trip ID = 1 tồn tại, chưa khởi hành và ghế J05 đang trống.', 'Người dùng hợp lệ đã được khởi tạo.', 'Request chứa đồng thời một ghế hợp lệ và một giá trị ghế rỗng.', 'Repository được mock để service vẫn có thể lưu booking nếu không validate toàn bộ danh sách.', 'Trip ID = 1.', 'seatNumbers = ["J05", ""].', 'paymentMethod = "CASH".', 'User hợp lệ.', 'Test Conditions: Kiểm tra lỗ hổng validation khi danh sách ghế trộn giữa dữ liệu hợp lệ và dữ liệu rỗng; toàn bộ request phải bị từ chối nếu có ít nhất một phần tử sai.', 'Gọi createBooking() với request có danh sách ghế gồm J05 và chuỗi rỗng.', 'Theo dõi việc service xử lý toàn bộ danh sách ghế và lưu booking detail.', 'Service vẫn tạo booking và chấp nhận cả phần tử ghế rỗng, không loại bỏ toàn bộ request.', 'Hệ thống phải từ chối toàn bộ request do danh sách ghế chứa dữ liệu không hợp lệ.'),
        }
        v = mapping[name]
        return pack([v[0], v[1], v[2], v[3]], [v[4], v[5], v[6], v[7]], v[8], fail_steps(v[9], v[10], v[11], v[12]))

    mapping = {
        'TC-21': ('Trip ID = 1 tồn tại, chưa khởi hành và ghế E01 đang trống.', 'Người dùng hợp lệ đã được khởi tạo.', 'EmailService được mock để ném exception khi gửi email xác nhận.', 'BookingRepository vẫn được mock để lưu booking thành công.', 'Trip ID = 1.', 'seatNumbers = ["E01"].', 'paymentMethod = "CASH".', 'EmailService.sendEmail(...) ném RuntimeException("SMTP unavailable").', 'Test Conditions: Chạy kiểm thử lifecycle nhằm xác minh việc gửi email lỗi không làm rollback giao dịch đặt vé tiền mặt đã được lưu thành công.', 'Thiết lập EmailService ném exception, đồng thời giữ các mock còn lại ở trạng thái happy path.', 'Theo dõi trạng thái booking trả về và số lần save(), sendEmail() được gọi.', 'Booking vẫn được tạo thành công với trạng thái CONFIRMED; lỗi email chỉ được ghi log, testcase PASS.'),
        'TC-22': ('Hàm checkAndCancelExpiredBookings() đã sẵn sàng để gọi trực tiếp.', 'Repository được mock trả về cancelledCount = 0.', 'Không có booking PENDING nào quá 15 phút trong kịch bản kiểm thử.', 'Môi trường test cho phép verify số lần repository được gọi.', 'cancelExpiredPendingBookings(any(LocalDateTime)) trả về 0.', 'Không có dữ liệu booking quá hạn trong repository.', 'Mốc thời gian hiện tại của scheduler.', 'Không có email hay thao tác phụ nào khác liên quan.', 'Test Conditions: Chạy scheduled task trong tình huống không có dữ liệu quá hạn; mục tiêu là bảo đảm job vẫn chạy ổn định và không phát sinh lỗi.', 'Mock bookingRepository.cancelExpiredPendingBookings(...) trả về 0.', 'Theo dõi việc repository được gọi và phản ứng của service.', 'Scheduled task chạy thành công, không hủy bản ghi nào và không phát sinh lỗi; testcase PASS.'),
        'TC-23': ('Hàm checkAndCancelExpiredBookings() đã được khởi tạo đầy đủ dependency.', 'BookingRepository được mock để ném RuntimeException khi xử lý cron job.', 'Môi trường test cho phép quan sát việc service có nuốt exception hay không.', 'Không yêu cầu dữ liệu booking cụ thể vì lỗi phát sinh ngay tại repository.', 'cancelExpiredPendingBookings(any(LocalDateTime)) ném RuntimeException("Database error").', 'Không có dữ liệu đầu vào từ người dùng.', 'Cron job được gọi trực tiếp từ unit test.', 'Service phải bắt lỗi và tiếp tục hoạt động.', 'Test Conditions: Kiểm tra khả năng chịu lỗi của scheduled task; dù repository phát sinh exception, hệ thống không được crash hoặc làm dừng luồng xử lý.', 'Mock bookingRepository.cancelExpiredPendingBookings(...) ném RuntimeException.', 'Theo dõi phản ứng của service khi repository ném lỗi.', 'Exception được service xử lý nội bộ, hệ thống không crash và testcase PASS.'),
        'TC-24': ('Trip ID = 1 tồn tại, chưa khởi hành và có giá vé hợp lệ.', 'Hai ghế F01 và F02 đều đang trống.', 'Chuẩn bị 2 người dùng khác nhau để gửi request đồng thời.', 'Repository được mock để ban đầu không phát hiện ghế trùng và theo dõi save() được gọi bao nhiêu lần.', 'User A đặt seatNumbers = ["F01"].', 'User B đặt seatNumbers = ["F02"].', 'Trip ID = 1 cho cả hai request.', 'paymentMethod = "CASH" cho cả hai request.', 'Test Conditions: Chạy kiểm thử đồng thời nhưng trên hai ghế khác nhau; xác minh hệ thống vẫn xử lý đúng khi không có tranh chấp tài nguyên ghế.', 'Tạo 2 luồng đồng thời gọi createBooking() với 2 request khác nhau cho F01 và F02.', 'Kích hoạt hai luồng chạy gần như cùng lúc và chờ cả hai hoàn tất.', 'Cả hai booking đều được lưu thành công, bookingRepository.save() được gọi 2 lần; testcase PASS.'),
        'TC-25': ('Booking ID = 101 tồn tại trong hệ thống.', 'Booking ở trạng thái PENDING và thuộc về User A.', 'Trip liên kết với booking chưa khởi hành.', 'BookingRepository được mock để findById() và save() hoạt động bình thường.', 'bookingId = 101.', 'user = User A (chủ sở hữu booking).', 'booking.status = "PENDING".', 'trip.departureTime > now.', 'Test Conditions: Kiểm tra luồng người dùng tự hủy vé hợp lệ; chỉ chủ sở hữu booking mới được phép hủy khi chuyến đi chưa khởi hành.', 'Mock bookingRepository.findById(101) trả về booking hợp lệ thuộc về User A.', 'Theo dõi trạng thái booking sau khi service xử lý và lưu lại dữ liệu.', 'Booking được cập nhật sang CANCELLED và save() được gọi đúng 1 lần; testcase PASS.'),
        'TC-26': ('Booking ID = 102 tồn tại trong hệ thống.', 'Booking thuộc về User A nhưng request hủy được gửi bởi User B.', 'BookingRepository được mock để trả về booking khi findById().', 'Trip của booking chưa khởi hành để chỉ kiểm tra ràng buộc quyền sở hữu.', 'bookingId = 102.', 'user gửi yêu cầu = User B.', 'booking.owner = User A.', 'booking.status = "PENDING".', 'Test Conditions: Kiểm tra luồng hủy vé sai chủ sở hữu; service phải chặn yêu cầu trước khi thay đổi trạng thái booking.', 'Mock bookingRepository.findById(102) trả về booking thuộc User A.', 'Gọi cancelBooking(102, User B) với user không phải chủ booking.', 'Service ném RuntimeException đúng như mong đợi, booking không được hủy; testcase PASS.'),
        'TC-27': ('Booking ID = 103 tồn tại trong hệ thống.', 'Booking thuộc về đúng người dùng gửi yêu cầu.', 'Trip của booking đã khởi hành tại thời điểm test.', 'BookingRepository được mock để trả về booking hợp lệ khi findById().', 'bookingId = 103.', 'user = chủ sở hữu booking.', 'booking.status = "PENDING".', 'trip.departureTime < now.', 'Test Conditions: Kiểm tra ràng buộc thời gian khi hủy vé; dù đúng chủ sở hữu, người dùng không được hủy booking nếu chuyến đi đã khởi hành.', 'Mock bookingRepository.findById(103) trả về booking thuộc chuyến đi đã khởi hành.', 'Quan sát phản hồi của service khi so sánh thời gian khởi hành của trip.', 'Service ném RuntimeException vì chuyến đi đã khởi hành; testcase PASS.'),
        'TC-28': ('Booking ID = 104 tồn tại trong hệ thống.', 'Booking đang ở trạng thái CONFIRMED.', 'BookingRepository được mock để findById() và save() hoạt động bình thường.', 'Không yêu cầu kiểm tra quyền sở hữu vì đây là luồng admin hủy booking.', 'bookingId = 104.', 'booking.status = "CONFIRMED".', 'Admin gọi cancelBookingByAdmin().', 'BookingRepository.save(...) trả về booking sau cập nhật.', 'Test Conditions: Kiểm tra luồng quản trị viên hủy booking đã xác nhận; hệ thống phải cho phép admin chuyển trạng thái booking sang CANCELLED.', 'Mock bookingRepository.findById(104) trả về booking đang CONFIRMED.', 'Theo dõi trạng thái booking sau khi service xử lý và lưu lại dữ liệu.', 'Booking được cập nhật thành CANCELLED và admin hủy booking thành công; testcase PASS.'),
        'TC-29': ('Booking đang tồn tại với trạng thái PENDING.', 'Booking có momoOrderId = ORD-01 và momoRequestId = REQ-01.', 'BookingRepository được mock để tìm thấy booking theo cặp mã giao dịch.', 'EmailService được mock để gửi email xác nhận thành công.', 'momoOrderId = "ORD-01".', 'momoRequestId = "REQ-01".', 'booking.status ban đầu = "PENDING".', 'bookingRepository.findAllWithDetails() trả về booking để phục vụ gửi email.', 'Test Conditions: Kiểm tra luồng callback xác nhận thanh toán MoMo thành công; booking PENDING phải được cập nhật sang CONFIRMED và kích hoạt gửi email xác nhận.', 'Mock bookingRepository.findByMomoOrderIdAndMomoRequestId(...) trả về booking PENDING hợp lệ.', 'Theo dõi việc service cập nhật trạng thái, save() booking và gọi EmailService.', 'Booking được chuyển sang CONFIRMED, save() và sendEmail() đều được gọi đúng như mong đợi; testcase PASS.'),
        'TC-30': ('Không tồn tại booking nào khớp với bộ mã giao dịch được cung cấp.', 'BookingRepository được mock để findByMomoOrderIdAndMomoRequestId() trả về Optional.empty().', 'EmailService được mock để kiểm tra không bị gọi ngoài ý muốn.', 'Môi trường unit test cho phép verify save() và sendEmail() không xảy ra.', 'momoOrderId = "ORD-404".', 'momoRequestId = "REQ-404".', 'bookingRepository.findByMomoOrderIdAndMomoRequestId(...) trả về Optional.empty().', 'Không có booking PENDING tương ứng trong dữ liệu mock.', 'Test Conditions: Kiểm tra callback thanh toán MoMo khi không tìm thấy booking phù hợp; service phải kết thúc an toàn và không thực hiện lưu dữ liệu hoặc gửi email.', 'Mock bookingRepository.findByMomoOrderIdAndMomoRequestId(...) trả về Optional.empty().', 'Theo dõi việc service có gọi save() hoặc sendEmail() hay không.', 'Không có booking nào được cập nhật, save() và sendEmail() không bị gọi; testcase PASS.'),
    }
    v = mapping[name]
    return pack([v[0], v[1], v[2], v[3]], [v[4], v[5], v[6], v[7]], v[8], pass_steps(v[9], v[10], v[11]))


def copy_style(ws, target_row, source_row):
    for col in range(1, 7):
        ws.cell(target_row, col)._style = copy(template.cell(source_row, col)._style)


def rebuild(ws, meta):
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= 8 or rng.max_row >= 8:
            ws.unmerge_cells(str(rng))
    if ws.max_row >= 8:
        ws.delete_rows(8, ws.max_row - 7)

    heights = {8: 33.6, 9: 33.6, 10: 33.6, 11: 33.6, 12: 33.6, 13: 18, 14: 18, 15: 28, 16: 42, 17: 42, 18: 42, 19: 42, 20: 42}
    for row in range(8, 21):
        src = 8 if row == 8 else row if row in (9, 10, 11, 12, 13, 14, 15, 16) else (17 if row % 2 == 1 else 18)
        copy_style(ws, row, src)
        ws.row_dimensions[row].height = heights[row]

    for row in (13, 14):
        for col in range(1, 7):
            ws.cell(row, col).value = None

    ws.cell(8, 1).value = 'S #'
    ws.cell(8, 2).value = 'Prerequisites:'
    ws.cell(8, 4).value = 'S #'
    ws.cell(8, 5).value = 'Test Data Requirement'

    for i in range(4):
        ws.cell(9 + i, 1).value = i + 1
        ws.cell(9 + i, 2).value = meta['pr'][i]
        ws.cell(9 + i, 4).value = i + 1
        ws.cell(9 + i, 5).value = meta['dt'][i]

    ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=5)
    ws.cell(15, 1).value = meta['cond']

    headers = ['Step #', 'Step Details', 'Expected Results', 'Actual Results', 'Pass / Fail / Not executed / Suspended']
    for col, value in enumerate(headers, 1):
        ws.cell(16, col).value = value

    for idx, (detail, expected, actual, status) in enumerate(meta['steps'], start=17):
        ws.cell(idx, 1).value = idx - 16
        ws.cell(idx, 2).value = detail
        ws.cell(idx, 3).value = expected
        ws.cell(idx, 4).value = actual
        ws.cell(idx, 5).value = status


for sheet in wb.sheetnames[1:]:
    rebuild(wb[sheet], data_for(sheet))

wb.save(DST)
print(DST)
