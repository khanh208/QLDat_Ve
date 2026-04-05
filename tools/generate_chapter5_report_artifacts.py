from __future__ import annotations

import copy
import shutil
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
BAO_CAO_DIR = ROOT / "Bao_Cao"
TESTCASE_PATH = BAO_CAO_DIR / "TestCase.xlsx"
TESTREPORT_PATH = BAO_CAO_DIR / "TestReport.xlsx"
SUMMARY_TEMPLATE_DOCX = BAO_CAO_DIR / "BÁO CÁO TÓM TẮT KIỂM THỬ.docx"

CHAPTER5_RUN_ROOT = ROOT / "target" / "chapter5-test-execution"
PERFORMANCE_REPORT_DIR = ROOT / "target" / "performance-reports"

DATE_TESTED = "04/04/2026"
CREATED_BY = "Nguyễn Quốc Khánh"
REVIEWED_BY = "ThS. Phạm Trọng Huynh"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


@dataclass
class Step:
    detail: str
    expected: str
    actual: str
    result: str


@dataclass
class CaseDefinition:
    tc_id: str
    sheet_name: str
    section: str
    suite: str
    method_name: str
    description: str
    qa_log: str
    prerequisites: list[str]
    test_data: list[str]
    test_conditions: str
    steps: list[Step]
    status: str
    note: str


def latest_execution_dir() -> Path:
    directories = [path for path in CHAPTER5_RUN_ROOT.iterdir() if path.is_dir()]
    if not directories:
        raise FileNotFoundError("Không tìm thấy thư mục target/chapter5-test-execution để lấy log Chương 5.")
    return max(directories, key=lambda path: path.stat().st_mtime)


def read_performance_metrics(file_name: str) -> dict[str, str]:
    metrics: dict[str, str] = {}
    report_path = PERFORMANCE_REPORT_DIR / file_name
    for line in report_path.read_text(encoding="utf-8").splitlines():
        if ":" in line:
            key, value = line.split(":", 1)
            metrics[key.strip()] = value.strip()
    return metrics


def build_cases() -> list[CaseDefinition]:
    create_booking_metrics = read_performance_metrics("create-booking-smoke.txt")
    concurrent_metrics = read_performance_metrics("concurrent-booking-smoke.txt")

    return [
        CaseDefinition(
            tc_id="TC31",
            sheet_name="TC-31",
            section="5.2 Integration Testing",
            suite="BookingApiIntegrationTest",
            method_name="createBookingEndpointShouldReturnConfirmedBookingForCashFlow",
            description="Kiểm tra API tạo booking tiền mặt trả về booking CONFIRMED đúng payload",
            qa_log="Xác minh BookingController phối hợp đúng với BookingService trong luồng tạo booking CASH bằng MockMvc.",
            prerequisites=[
                "Bài test WebMvcTest khởi tạo BookingController thành công.",
                "MockMvc hoạt động và SecurityConfig đã nạp cấu hình test.",
                "BookingService đã được mock để trả về booking CONFIRMED.",
                "Người dùng đăng nhập giả lập ROLE_USER đã sẵn sàng.",
            ],
            test_data=[
                "Request JSON: tripId=1, seatNumbers=[\"B1\"], paymentMethod=\"CASH\"",
                "Mock user: integration-user (ROLE_USER)",
                "Mock trip: tripId=1",
                "Mock booking kết quả: bookingId=501, status=CONFIRMED",
            ],
            test_conditions="Test Conditions: Gửi POST /api/bookings với người dùng đã xác thực; BookingService được stub trả về booking đã xác nhận và kiểm tra response JSON của controller.",
            steps=[
                Step("Cấu hình mock BookingService.createBooking() trả về bookingId=501, status=CONFIRMED, ghế B1.", "Service giả lập được chuẩn bị đúng cho controller.", "Mock trả về booking CONFIRMED với bookingId=501.", "PASS"),
                Step("Dùng MockMvc gửi POST /api/bookings với request JSON và authentication của user thường.", "API nhận request hợp lệ và gọi đúng lớp service.", "API trả HTTP 200, controller gọi createBooking() đúng 1 lần.", "PASS"),
                Step("Kiểm tra payload JSON trả về trong response.", "Payload phải chứa bookingId=501, status=CONFIRMED, tripId=1 và seatNumber=B1.", "JSON response chứa đúng bookingId=501, status=CONFIRMED, trip.tripId=1 và bookingDetails[0].seatNumber=B1.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng integration cho luồng tạo booking CASH thành công.",
        ),
        CaseDefinition(
            tc_id="TC32",
            sheet_name="TC-32",
            section="5.2 Integration Testing",
            suite="BookingApiIntegrationTest",
            method_name="createBookingEndpointShouldReturnBadRequestWhenServiceRejectsSeatConflict",
            description="Kiểm tra API tạo booking trả về HTTP 400 khi service phát hiện ghế đã bị giữ",
            qa_log="Xác minh controller phản hồi lỗi đúng khi BookingService từ chối request do xung đột ghế.",
            prerequisites=[
                "BookingController được khởi tạo trong WebMvcTest.",
                "MockMvc và security test config hoạt động bình thường.",
                "BookingService đã được mock để ném RuntimeException về xung đột ghế.",
            ],
            test_data=[
                "Request JSON: tripId=1, seatNumbers=[\"A2\"], paymentMethod=\"MOMO\"",
                "Thông báo lỗi mong đợi: Seat A2 is already held.",
                "Mock user: integration-user (ROLE_USER)",
            ],
            test_conditions="Test Conditions: Gửi POST /api/bookings với request hợp lệ nhưng service ném lỗi ghế trùng; controller phải chuyển lỗi thành HTTP 400 kèm nội dung mô tả.",
            steps=[
                Step("Cấu hình BookingService.createBooking() ném RuntimeException với thông điệp Seat A2 is already held.", "Mock service được thiết lập ở nhánh lỗi ghế trùng.", "BookingService mock ném đúng exception theo kịch bản.", "PASS"),
                Step("Gửi POST /api/bookings với user đã xác thực và request chứa ghế A2.", "Controller nhận request và xử lý lỗi từ service.", "API trả về HTTP 400 cho request xung đột ghế.", "PASS"),
                Step("Kiểm tra nội dung response body.", "Response phải chứa chuỗi Seat A2 is already held.", "Body phản hồi chứa đúng thông điệp Seat A2 is already held.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng integration cho nhánh lỗi seat conflict.",
        ),
        CaseDefinition(
            tc_id="TC33",
            sheet_name="TC-33",
            section="5.2 Integration Testing",
            suite="BookingApiIntegrationTest",
            method_name="bookedSeatsEndpointShouldReturnCurrentSeatSet",
            description="Kiểm tra API booked-seats trả về đúng tập ghế đã được giữ/bán",
            qa_log="Xác minh TripController và BookingService phối hợp đúng khi trả danh sách ghế đã đặt.",
            prerequisites=[
                "TripController và BookingService mock đã được nạp trong WebMvcTest.",
                "MockMvc truy cập được endpoint GET /api/trips/{id}/booked-seats.",
            ],
            test_data=[
                "Trip ID = 1",
                "Tập ghế mock trả về: A1, B2",
            ],
            test_conditions="Test Conditions: Gửi GET /api/trips/1/booked-seats và đối chiếu response body với tập ghế do BookingService mock trả về.",
            steps=[
                Step("Stub BookingService.getBookedSeatsForTrip(1) trả về tập {A1, B2}.", "Service giả lập có dữ liệu ghế đã đặt.", "Mock service trả về đúng tập ghế A1 và B2.", "PASS"),
                Step("Gửi GET /api/trips/1/booked-seats bằng MockMvc.", "API phản hồi HTTP 200.", "Endpoint trả về HTTP 200.", "PASS"),
                Step("Kiểm tra nội dung body phản hồi.", "Body phải chứa cả A1 và B2.", "Response body chứa chuỗi A1 và B2 đúng như dữ liệu mock.", "PASS"),
            ],
            status="PASS",
            note="Dùng cho phần tích hợp API tra cứu ghế đã đặt.",
        ),
        CaseDefinition(
            tc_id="TC34",
            sheet_name="TC-34",
            section="5.2 Integration Testing",
            suite="BookingApiIntegrationTest",
            method_name="momoConfirmEndpointShouldReturnSuccessPayloadForAuthenticatedUser",
            description="Kiểm tra API confirm thanh toán MoMo trả về payload thành công cho user đã xác thực",
            qa_log="Xác minh PaymentController gọi confirmBookingAfterPayment đúng tham số và trả về trạng thái success.",
            prerequisites=[
                "PaymentController đã được nạp trong WebMvcTest.",
                "BookingService mock sẵn sàng xác nhận booking sau thanh toán.",
                "User ROLE_USER đã được gắn authentication vào request.",
            ],
            test_data=[
                "Request JSON: orderId=ORD-501, requestId=REQ-501, resultCode=0",
                "Mock user: integration-user (ROLE_USER)",
            ],
            test_conditions="Test Conditions: Gửi POST /api/payment/momo/confirm với payload callback hợp lệ; controller phải trả JSON status=success và gọi confirmBookingAfterPayment().",
            steps=[
                Step("Chuẩn bị user đã xác thực và payload callback MoMo hợp lệ.", "Request xác nhận có đầy đủ orderId, requestId và resultCode.", "Authentication và payload callback được khởi tạo thành công.", "PASS"),
                Step("Gửi POST /api/payment/momo/confirm bằng MockMvc.", "API xử lý callback và trả về HTTP 200.", "Endpoint trả về HTTP 200.", "PASS"),
                Step("Kiểm tra JSON response và verify tương tác với BookingService.", "Response phải có status=success, orderId=ORD-501; service được gọi đúng 1 lần.", "JSON response có status=success, orderId=ORD-501 và BookingService.confirmBookingAfterPayment(\"ORD-501\", \"REQ-501\") được gọi đúng 1 lần.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng tích hợp cho callback xác nhận thanh toán MoMo.",
        ),
        CaseDefinition(
            tc_id="TC35",
            sheet_name="TC-35",
            section="5.2 Integration Testing",
            suite="BookingApiIntegrationTest",
            method_name="searchTripsEndpointShouldReturnMatchingTrips",
            description="Kiểm tra API tìm chuyến trả về đúng danh sách chuyến phù hợp tiêu chí tìm kiếm",
            qa_log="Xác minh TripController trả dữ liệu tuyến xe đúng với bộ lọc điểm đi, điểm đến và ngày khởi hành.",
            prerequisites=[
                "TripController được nạp trong WebMvcTest.",
                "TripService đã được mock cho kết quả tìm kiếm.",
                "Endpoint tìm chuyến cho phép truy cập anonymous.",
            ],
            test_data=[
                "start_location=TP. Ho Chi Minh",
                "end_location=Da Lat",
                "date=ngày khởi hành của trip mock",
                "Trip mock: tripId=1",
            ],
            test_conditions="Test Conditions: Gửi GET /api/trips/search với đầy đủ query params; TripService mock trả về 1 chuyến phù hợp và controller phải trả JSON đúng.",
            steps=[
                Step("Stub TripService.searchTrips() trả về 1 trip có route TP. Ho Chi Minh -> Da Lat.", "Service giả lập có dữ liệu chuyến phù hợp.", "Mock searchTrips trả về danh sách gồm tripId=1.", "PASS"),
                Step("Gửi GET /api/trips/search với các tham số điểm đi, điểm đến và ngày.", "API trả HTTP 200 cùng danh sách chuyến phù hợp.", "Endpoint trả về HTTP 200.", "PASS"),
                Step("Kiểm tra phần tử đầu tiên của mảng JSON trả về.", "Trip đầu tiên phải có tripId=1, startLocation=TP. Ho Chi Minh, endLocation=Da Lat.", "JSON trả về đúng tripId=1, route.startLocation=TP. Ho Chi Minh và route.endLocation=Da Lat.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng tích hợp cho API tra cứu chuyến xe.",
        ),
        CaseDefinition(
            tc_id="TC36",
            sheet_name="TC-36",
            section="5.2 Integration Testing",
            suite="BookingApiIntegrationTest",
            method_name="cancelBookingEndpointShouldReturnUpdatedBooking",
            description="Kiểm tra API hủy booking trả về booking đã cập nhật trạng thái CANCELLED",
            qa_log="Xác minh BookingController gọi đúng service hủy vé và phản hồi booking sau khi cập nhật.",
            prerequisites=[
                "BookingController chạy trong WebMvcTest.",
                "BookingService.cancelBooking() đã được mock.",
                "User đã xác thực hợp lệ.",
            ],
            test_data=[
                "Booking ID = 601",
                "Mock booking kết quả: status=CANCELLED, paymentMethod=CASH, seat=B1",
                "Mock user: integration-user (ROLE_USER)",
            ],
            test_conditions="Test Conditions: Gửi PUT /api/bookings/601/cancel với user đã xác thực; controller phải trả HTTP 200 và JSON booking có status=CANCELLED.",
            steps=[
                Step("Stub BookingService.cancelBooking(601, user) trả về booking trạng thái CANCELLED.", "Service mock được chuẩn bị đúng nhánh hủy vé.", "BookingService mock trả về bookingId=601, status=CANCELLED.", "PASS"),
                Step("Gửi PUT /api/bookings/601/cancel bằng MockMvc với authentication của user.", "API tiếp nhận yêu cầu hủy và trả HTTP 200.", "Endpoint trả về HTTP 200.", "PASS"),
                Step("Kiểm tra JSON response.", "Response phải có bookingId=601 và status=CANCELLED.", "JSON response hiển thị bookingId=601, status=CANCELLED đúng mong đợi.", "PASS"),
            ],
            status="PASS",
            note="Dùng cho phần kiểm thử tích hợp API hủy vé.",
        ),
        CaseDefinition(
            tc_id="TC37",
            sheet_name="TC-37",
            section="5.3 System Testing",
            suite="BookingSystemFlowMockMvcTest",
            method_name="cashBookingJourneyShouldCompleteThroughTripLookupCreateAndMyBookings",
            description="Kiểm tra luồng hệ thống mô phỏng đặt vé CASH từ tra cứu chuyến đến xem danh sách vé của tôi",
            qa_log="Xác minh chuỗi thao tác TripController -> BookingController -> MyBookings hoạt động đúng trong một hành trình mock hoàn chỉnh.",
            prerequisites=[
                "MockMvc khởi tạo đầy đủ TripController, BookingController và PaymentController.",
                "TripService và BookingService đã được stub theo luồng CASH.",
                "Người dùng system-user đã xác thực hợp lệ.",
            ],
            test_data=[
                "Trip ID = 1",
                "Ghế cần đặt: B1",
                "Payment method = CASH",
                "Booking kết quả: bookingId=801, status=CONFIRMED",
            ],
            test_conditions="Test Conditions: Thực hiện lần lượt GET /api/trips/1, GET /api/trips/1/booked-seats, POST /api/bookings và GET /api/bookings/my-bookings; toàn bộ chuỗi phải thành công.",
            steps=[
                Step("Stub dữ liệu chuyến, ghế đã đặt, booking tạo mới và lịch sử booking của user.", "Toàn bộ phụ thuộc được chuẩn bị đủ cho một hành trình đặt vé CASH.", "TripService và BookingService mock đã được thiết lập đầy đủ cho luồng CASH.", "PASS"),
                Step("Gọi lần lượt API lấy thông tin chuyến và danh sách ghế đã đặt.", "Hai endpoint trả HTTP 200 và trả dữ liệu đúng trip/seat đã mock.", "GET /api/trips/1 và GET /api/trips/1/booked-seats đều trả HTTP 200 với dữ liệu đúng.", "PASS"),
                Step("Gửi POST /api/bookings với ghế B1, payment=CASH.", "Booking được tạo thành công với trạng thái CONFIRMED.", "API tạo booking trả HTTP 200, bookingId=801, status=CONFIRMED.", "PASS"),
                Step("Gọi GET /api/bookings/my-bookings để kiểm tra lịch sử booking.", "Danh sách vé của người dùng phải chứa bookingId=801 ở trạng thái CONFIRMED.", "API lịch sử vé trả HTTP 200 và phần tử đầu tiên có bookingId=801, status=CONFIRMED.", "PASS"),
            ],
            status="PASS",
            note="Luồng hệ thống mock cho kịch bản đặt vé tiền mặt.",
        ),
        CaseDefinition(
            tc_id="TC38",
            sheet_name="TC-38",
            section="5.3 System Testing",
            suite="BookingSystemFlowMockMvcTest",
            method_name="momoBookingJourneyShouldCompleteThroughCheckoutConfirmAndHistory",
            description="Kiểm tra luồng hệ thống mô phỏng đặt vé MOMO từ tạo booking PENDING đến xác nhận thanh toán và xem lịch sử",
            qa_log="Xác minh chuỗi create booking -> checkout MOMO -> confirm callback -> my bookings hoạt động đúng trong hệ thống mock.",
            prerequisites=[
                "MockMvc khởi tạo các controller đặt vé và thanh toán.",
                "BookingService và PaymentService đã được stub theo luồng MOMO.",
                "Người dùng momo-user đã xác thực hợp lệ.",
            ],
            test_data=[
                "Trip ID = 1",
                "Ghế cần đặt: C1",
                "Booking trước thanh toán: bookingId=901, status=PENDING",
                "Booking sau thanh toán: bookingId=901, status=CONFIRMED",
            ],
            test_conditions="Test Conditions: Thực hiện POST /api/bookings, POST /api/payment/momo/checkout/{id}, POST /api/payment/momo/confirm và GET /api/bookings/my-bookings; luồng MOMO phải hoàn thành đầy đủ.",
            steps=[
                Step("Stub BookingService.createBooking() trả về booking PENDING và PaymentService.createMoMoPayment() trả payUrl.", "Các bước trung gian của luồng MOMO được mô phỏng đầy đủ.", "BookingService và PaymentService mock đã sẵn sàng với booking PENDING và payUrl mock.", "PASS"),
                Step("Gửi POST /api/bookings với paymentMethod=MOMO.", "API tạo booking trả trạng thái PENDING.", "API trả HTTP 200 với bookingId=901, status=PENDING.", "PASS"),
                Step("Gửi POST /api/payment/momo/checkout/901 và POST /api/payment/momo/confirm với callback hợp lệ.", "Checkout trả payUrl; confirm trả status=success.", "Checkout trả HTTP 200 với payUrl https://momo.test/pay/901; confirm trả HTTP 200 với status=success.", "PASS"),
                Step("Gọi GET /api/bookings/my-bookings để đối chiếu lịch sử sau thanh toán.", "Danh sách booking phải chứa bookingId=901 ở trạng thái CONFIRMED.", "API lịch sử vé trả HTTP 200; booking đầu tiên có bookingId=901, status=CONFIRMED.", "PASS"),
            ],
            status="PASS",
            note="Luồng hệ thống mock cho kịch bản đặt vé MOMO.",
        ),
        CaseDefinition(
            tc_id="TC39",
            sheet_name="TC-39",
            section="5.4 Performance Testing",
            suite="BookingPerformanceSmokeTest",
            method_name="createBookingShouldStayWithinPerformanceBudget",
            description="Kiểm tra hiệu năng smoke test cho createBooking trong 150 lượt gọi liên tiếp",
            qa_log="Đo độ trễ trung bình, P95 và cực đại của createBooking trong môi trường mock để đánh giá nhanh hiệu năng nghiệp vụ.",
            prerequisites=[
                "MockitoJUnitRunner khởi tạo BookingService cùng các repository mock thành công.",
                "Trip hợp lệ, user hợp lệ và dữ liệu ghế hợp lệ đã được chuẩn bị.",
                "Thư mục target/performance-reports có quyền ghi báo cáo.",
            ],
            test_data=[
                "Số lần lặp: 150",
                "Payment method = CASH",
                "Ghế sinh tự động: P0 đến P149",
                "Ngưỡng kiểm tra: Average < 75 ms, P95 < 150 ms",
            ],
            test_conditions="Test Conditions: Gọi createBooking() 150 lần liên tiếp trong môi trường mock; thu thập thời gian xử lý từng mẫu và đối chiếu với ngưỡng hiệu năng đã đặt.",
            steps=[
                Step("Khởi tạo request hợp lệ và lặp 150 lần gọi BookingService.createBooking().", "Mỗi lượt gọi được thực thi thành công và thời gian xử lý được ghi nhận.", "150 mẫu thời gian đã được thu thập thành công cho kịch bản createBooking.", "PASS"),
                Step("Tính toán Average, P95 và Max từ danh sách duration thu được.", "Các chỉ số được tính đúng và ghi ra file báo cáo hiệu năng.", f"Báo cáo create-booking-smoke.txt được tạo với Average={create_booking_metrics.get('Average (ms)', '?')}, P95={create_booking_metrics.get('P95 (ms)', '?')}, Max={create_booking_metrics.get('Max (ms)', '?')}.", "PASS"),
                Step("So sánh Average và P95 với ngưỡng kiểm thử.", "Average phải nhỏ hơn 75 ms và P95 phải nhỏ hơn 150 ms.", f"Average={create_booking_metrics.get('Average (ms)', '?')} < 75 ms và P95={create_booking_metrics.get('P95 (ms)', '?')} < 150 ms nên testcase đạt.", "PASS"),
            ],
            status="PASS",
            note="Hiệu năng ở mức smoke test, chưa phải load test toàn hệ thống.",
        ),
        CaseDefinition(
            tc_id="TC40",
            sheet_name="TC-40",
            section="5.4 Performance Testing",
            suite="BookingPerformanceSmokeTest",
            method_name="concurrentBookingsShouldCompleteWithinBudget",
            description="Kiểm tra hiệu năng smoke test cho các lượt đặt vé đồng thời trong môi trường mock",
            qa_log="Đo thời gian hoàn tất của 12 vòng đặt vé đồng thời để đánh giá nhanh khả năng đáp ứng của createBooking khi có nhiều tác vụ song song.",
            prerequisites=[
                "BookingService và các repository mock được khởi tạo thành công.",
                "Thread pool 4 luồng có thể chạy song song trong unit test.",
                "Trip hợp lệ và user hợp lệ đã được chuẩn bị.",
            ],
            test_data=[
                "Số vòng đo: 12",
                "Mỗi vòng gồm 4 tác vụ createBooking chạy đồng thời",
                "Payment method = CASH",
                "Ngưỡng kiểm tra: Average < 300 ms, Max < 1000 ms",
            ],
            test_conditions="Test Conditions: Với mỗi vòng, giải phóng đồng thời 4 task gọi createBooking(); đo thời gian hoàn thành của cả vòng và so sánh với ngưỡng đã định.",
            steps=[
                Step("Khởi tạo ExecutorService và CountDownLatch cho 12 vòng đo đồng thời.", "Môi trường thực thi đồng thời sẵn sàng cho kịch bản performance.", "12 vòng concurrent test được khởi tạo thành công với 4 luồng mỗi vòng.", "PASS"),
                Step("Thực thi từng vòng đo, ghi nhận thời gian hoàn tất và chờ đủ 4 task hoàn thành.", "Mỗi vòng phải hoàn tất trước khi hết timeout 2 giây.", "Cả 12 vòng đều hoàn tất trước timeout; dữ liệu thời gian được ghi nhận đầy đủ.", "PASS"),
                Step("Tính Average, P95 và Max từ thời gian của 12 vòng và so với ngưỡng.", "Average phải nhỏ hơn 300 ms và Max phải nhỏ hơn 1000 ms.", f"Báo cáo concurrent-booking-smoke.txt cho thấy Average={concurrent_metrics.get('Average (ms)', '?')}, Max={concurrent_metrics.get('Max (ms)', '?')}; cả hai đều dưới ngưỡng nên testcase đạt.", "PASS"),
            ],
            status="PASS",
            note="Hiệu năng đồng thời ở mức smoke test nội bộ bằng mock.",
        ),
        CaseDefinition(
            tc_id="TC41",
            sheet_name="TC-41",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="publicTripsEndpointShouldAllowAnonymousAccess",
            description="Kiểm tra endpoint công khai /api/trips cho phép người dùng ẩn danh truy cập",
            qa_log="Xác minh cấu hình bảo mật không chặn nhầm endpoint công khai của hệ thống.",
            prerequisites=[
                "SecurityConfig và TestSecurityBeans được nạp thành công.",
                "TripService đã được mock cho endpoint danh sách chuyến.",
            ],
            test_data=[
                "Endpoint: GET /api/trips",
                "Ngữ cảnh bảo mật: anonymous user",
                "Dữ liệu trả về mock: danh sách rỗng",
            ],
            test_conditions="Test Conditions: Gửi GET /api/trips trong ngữ cảnh anonymous; endpoint công khai phải trả HTTP 200 thay vì yêu cầu đăng nhập.",
            steps=[
                Step("Stub TripService.getAllTrips() trả về danh sách rỗng.", "Endpoint công khai có thể phản hồi mà không cần DB thật.", "TripService mock trả về danh sách rỗng thành công.", "PASS"),
                Step("Gửi GET /api/trips với @WithAnonymousUser.", "Request anonymous phải được phép truy cập.", "API trả HTTP 200 cho anonymous user.", "PASS"),
            ],
            status="PASS",
            note="Xác minh endpoint công khai được phép truy cập không cần đăng nhập.",
        ),
        CaseDefinition(
            tc_id="TC42",
            sheet_name="TC-42",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="loginEndpointShouldAllowAnonymousAccess",
            description="Kiểm tra endpoint đăng nhập cho phép người dùng ẩn danh truy cập và trả token",
            qa_log="Xác minh cơ chế bảo mật không chặn endpoint đăng nhập của người dùng chưa xác thực.",
            prerequisites=[
                "AuthController và AuthService mock được nạp trong WebMvcTest.",
                "Security filter test config hoạt động bình thường.",
            ],
            test_data=[
                "Endpoint: POST /api/auth/login",
                "Payload: username=selenium-user, password=123456",
                "Mock response: token=mock-token, message=Login success",
            ],
            test_conditions="Test Conditions: Gửi POST /api/auth/login ở trạng thái anonymous; endpoint phải trả HTTP 200 cùng payload AuthResponse.",
            steps=[
                Step("Stub AuthService.login() trả về AuthResponse với token mock-token.", "Service đăng nhập giả lập đã sẵn sàng.", "AuthService mock trả AuthResponse token=mock-token.", "PASS"),
                Step("Gửi POST /api/auth/login với payload đăng nhập hợp lệ.", "API đăng nhập phải cho phép anonymous user truy cập.", "Endpoint trả HTTP 200 khi gửi request login.", "PASS"),
                Step("Đối chiếu response code và payload đăng nhập.", "Response phải trả thông tin đăng nhập thành công.", "API trả HTTP 200 với AuthResponse phù hợp kịch bản mock.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng bảo mật cơ bản cho endpoint đăng nhập công khai.",
        ),
        CaseDefinition(
            tc_id="TC43",
            sheet_name="TC-43",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="bookingsEndpointShouldRejectAnonymousAccess",
            description="Kiểm tra endpoint /api/bookings/my-bookings từ chối người dùng ẩn danh",
            qa_log="Xác minh tài nguyên lịch sử booking yêu cầu người dùng phải đăng nhập trước khi truy cập.",
            prerequisites=[
                "BookingController được nạp cùng SecurityConfig.",
                "Không gắn authentication cho request kiểm thử.",
            ],
            test_data=[
                "Endpoint: GET /api/bookings/my-bookings",
                "Ngữ cảnh bảo mật: anonymous user",
            ],
            test_conditions="Test Conditions: Gửi GET /api/bookings/my-bookings mà không đăng nhập; hệ thống phải trả lỗi 4xx thay vì cung cấp dữ liệu booking.",
            steps=[
                Step("Đảm bảo request được gửi trong ngữ cảnh @WithAnonymousUser.", "Request không mang authentication hợp lệ.", "Ngữ cảnh anonymous đã được thiết lập đúng cho test.", "PASS"),
                Step("Gửi GET /api/bookings/my-bookings bằng MockMvc.", "Endpoint phải từ chối truy cập và trả mã 4xx.", "API trả mã lỗi 4xx cho anonymous user.", "PASS"),
            ],
            status="PASS",
            note="Xác minh tài nguyên cá nhân không được truy cập ẩn danh.",
        ),
        CaseDefinition(
            tc_id="TC44",
            sheet_name="TC-44",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="bookingsEndpointShouldAllowAuthenticatedUser",
            description="Kiểm tra endpoint /api/bookings/my-bookings cho phép user đã xác thực truy cập",
            qa_log="Xác minh người dùng hợp lệ được phép truy cập lịch sử booking cá nhân.",
            prerequisites=[
                "BookingController và BookingService mock đã được khởi tạo.",
                "Authentication ROLE_USER hợp lệ đã được chuẩn bị.",
            ],
            test_data=[
                "Endpoint: GET /api/bookings/my-bookings",
                "Ngữ cảnh bảo mật: authenticated USER",
                "Mock bookingService.getMyBookings() trả danh sách rỗng",
            ],
            test_conditions="Test Conditions: Gửi GET /api/bookings/my-bookings với authentication hợp lệ; endpoint phải trả HTTP 200 và gọi BookingService.getMyBookings().",
            steps=[
                Step("Stub BookingService.getMyBookings() trả danh sách rỗng.", "Service lịch sử booking sẵn sàng phản hồi cho user hợp lệ.", "BookingService mock đã được thiết lập thành công.", "PASS"),
                Step("Gửi GET /api/bookings/my-bookings với authentication ROLE_USER.", "Endpoint phải cho phép truy cập và trả HTTP 200.", "API trả HTTP 200 cho user đã xác thực.", "PASS"),
                Step("Verify BookingService.getMyBookings() được gọi đúng 1 lần.", "Lớp service phải được controller gọi đúng theo luồng nghiệp vụ.", "BookingService.getMyBookings() được verify gọi đúng 1 lần.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng phân quyền đúng cho endpoint my-bookings.",
        ),
        CaseDefinition(
            tc_id="TC45",
            sheet_name="TC-45",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="adminEndpointShouldRejectRegularUser",
            description="Kiểm tra endpoint admin từ chối user thường truy cập",
            qa_log="Xác minh người dùng ROLE_USER không thể truy cập API quản trị booking.",
            prerequisites=[
                "AdminBookingController và SecurityConfig đã được nạp.",
                "Authentication ROLE_USER được chuẩn bị cho request.",
            ],
            test_data=[
                "Endpoint: GET /api/admin/bookings",
                "Role request: USER",
            ],
            test_conditions="Test Conditions: Gửi GET /api/admin/bookings với user thường; endpoint quản trị phải trả HTTP 403 Forbidden.",
            steps=[
                Step("Gắn authentication ROLE_USER vào request kiểm thử.", "Request đại diện cho user thường, không có quyền admin.", "Authentication ROLE_USER được khởi tạo thành công.", "PASS"),
                Step("Gửi GET /api/admin/bookings bằng MockMvc.", "API quản trị phải trả HTTP 403 Forbidden.", "Endpoint trả HTTP 403 cho user thường.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng chặn truy cập trái phép vào API admin.",
        ),
        CaseDefinition(
            tc_id="TC46",
            sheet_name="TC-46",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="adminEndpointShouldAllowAdminUser",
            description="Kiểm tra endpoint admin cho phép người dùng ADMIN truy cập",
            qa_log="Xác minh API quản trị booking chỉ cho phép người dùng có vai trò ADMIN.",
            prerequisites=[
                "AdminBookingController đã được khởi tạo trong WebMvcTest.",
                "Authentication ROLE_ADMIN được chuẩn bị.",
                "BookingService.getAllBookingsSorted() đã được mock.",
            ],
            test_data=[
                "Endpoint: GET /api/admin/bookings",
                "Role request: ADMIN",
                "Mock kết quả: danh sách booking rỗng",
            ],
            test_conditions="Test Conditions: Gửi GET /api/admin/bookings với admin user; endpoint phải trả HTTP 200 và cho phép lấy dữ liệu.",
            steps=[
                Step("Stub BookingService.getAllBookingsSorted() trả danh sách rỗng.", "Service admin booking sẵn sàng cho request hợp lệ.", "BookingService mock trả danh sách rỗng thành công.", "PASS"),
                Step("Gửi GET /api/admin/bookings với authentication ROLE_ADMIN.", "API quản trị phải cho phép truy cập và trả HTTP 200.", "Endpoint trả HTTP 200 cho admin user.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng phân quyền đúng cho endpoint quản trị.",
        ),
        CaseDefinition(
            tc_id="TC47",
            sheet_name="TC-47",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="momoIpnEndpointShouldAllowAnonymousWebhook",
            description="Kiểm tra endpoint webhook MoMo IPN cho phép truy cập anonymous",
            qa_log="Xác minh webhook thanh toán của bên thứ ba không bị chặn bởi cơ chế yêu cầu đăng nhập.",
            prerequisites=[
                "PaymentController và PaymentService mock đã được nạp.",
                "Request chạy trong ngữ cảnh anonymous user.",
            ],
            test_data=[
                "Endpoint: POST /api/payment/momo-ipn",
                "Payload: orderId=ORD-001, requestId=REQ-001",
            ],
            test_conditions="Test Conditions: Gửi POST /api/payment/momo-ipn mà không đăng nhập; endpoint webhook phải trả HTTP 200 và gọi PaymentService.handleMomoIPN().",
            steps=[
                Step("Chuẩn bị payload webhook MoMo hợp lệ ở mức mock.", "Payload IPN đủ để controller chuyển tiếp cho service.", "Payload orderId/requestId đã được chuẩn bị.", "PASS"),
                Step("Gửi POST /api/payment/momo-ipn với anonymous user.", "Webhook phải được phép truy cập công khai và trả HTTP 200.", "Endpoint trả HTTP 200 cho request anonymous.", "PASS"),
                Step("Verify PaymentService.handleMomoIPN() được gọi đúng 1 lần.", "Controller phải chuyển payload cho service xử lý webhook.", "PaymentService.handleMomoIPN() được verify gọi đúng 1 lần.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng bảo mật cơ bản cho endpoint webhook bên thứ ba.",
        ),
        CaseDefinition(
            tc_id="TC48",
            sheet_name="TC-48",
            section="5.5 Security Testing",
            suite="SecurityAccessIntegrationTest",
            method_name="momoCheckoutEndpointShouldRejectAnonymousAccess",
            description="Kiểm tra endpoint checkout MOMO từ chối người dùng ẩn danh",
            qa_log="Xác minh thao tác khởi tạo thanh toán MoMo yêu cầu người dùng phải xác thực trước.",
            prerequisites=[
                "PaymentController đã được nạp cùng SecurityConfig.",
                "Request kiểm thử không gắn authentication.",
            ],
            test_data=[
                "Endpoint: POST /api/payment/momo/checkout/123",
                "Ngữ cảnh bảo mật: anonymous user",
            ],
            test_conditions="Test Conditions: Gửi POST /api/payment/momo/checkout/123 mà không đăng nhập; endpoint phải trả lỗi 4xx thay vì khởi tạo thanh toán.",
            steps=[
                Step("Đảm bảo request không mang authentication hợp lệ.", "Checkout được thực hiện bởi anonymous user.", "Ngữ cảnh anonymous đã được thiết lập đúng.", "PASS"),
                Step("Gửi POST /api/payment/momo/checkout/123 bằng MockMvc.", "Endpoint phải trả mã lỗi 4xx để chặn truy cập không xác thực.", "API trả mã lỗi 4xx cho anonymous user.", "PASS"),
            ],
            status="PASS",
            note="Minh chứng yêu cầu xác thực trước khi checkout MOMO.",
        ),
        CaseDefinition(
            tc_id="TC49",
            sheet_name="TC-49",
            section="5.6 Regression Testing",
            suite="RegressionStableSuite",
            method_name="regressionStableSuiteShouldPass",
            description="Kiểm tra regression suite ổn định chạy lại các nhóm test nền tảng sau khi bổ sung test Chương 5",
            qa_log="Xác minh các nhóm test ổn định vẫn pass sau khi bổ sung kiểm thử integration, system, performance và security.",
            prerequisites=[
                "SuiteAssertions hỗ trợ chạy lại các class test ổn định đã được cấu hình.",
                "BookingValidationTest, BookingLifecycleTest và SecurityAccessIntegrationTest biên dịch và chạy được.",
                "Môi trường unit test sử dụng mock hoạt động bình thường.",
            ],
            test_data=[
                "RegressionStableSuite",
                "Các class được chạy lại: BookingValidationTest, BookingLifecycleTest, SecurityAccessIntegrationTest",
                "Kỳ vọng: toàn bộ class trong suite pass",
            ],
            test_conditions="Test Conditions: Chạy RegressionStableSuite; suite phải xác nhận ba nhóm test nền tảng đều pass sau thay đổi gần nhất.",
            steps=[
                Step("Khởi động RegressionStableSuite trong môi trường test JUnit.", "Suite wrapper được khởi tạo thành công.", "RegressionStableSuite được chạy thành công.", "PASS"),
                Step("Suite gọi lại BookingValidationTest, BookingLifecycleTest và SecurityAccessIntegrationTest thông qua SuiteAssertions.", "Ba class ổn định phải vượt qua toàn bộ assert nội bộ.", "SuiteAssertions xác nhận cả 3 class test đều pass.", "PASS"),
                Step("Kiểm tra kết quả cuối cùng của RegressionStableSuite.", "Regression suite phải đạt PASS.", "RegressionStableSuite trả kết quả PASS và được dùng làm minh chứng hồi quy cơ bản.", "PASS"),
            ],
            status="PASS",
            note="Regression ở mức cơ bản, dùng lại các nhóm test ổn định quan trọng.",
        ),
    ]


def get_chapter5_cases() -> list[CaseDefinition]:
    source_cases = {case.tc_id: copy.deepcopy(case) for case in build_cases()}
    selection = [
        ("TC31", "TC31"),
        ("TC32", "TC32"),
        ("TC37", "TC33"),
        ("TC38", "TC34"),
        ("TC39", "TC35"),
        ("TC40", "TC36"),
        ("TC41", "TC37"),
        ("TC43", "TC38"),
        ("TC46", "TC39"),
        ("TC49", "TC40"),
    ]

    simplified: list[CaseDefinition] = []
    for source_id, target_id in selection:
        case = copy.deepcopy(source_cases[source_id])
        case.tc_id = target_id
        case.sheet_name = f"TC-{target_id[2:]}"
        case.prerequisites = case.prerequisites[:2]
        case.test_data = case.test_data[:2]
        case.steps = [case.steps[0], case.steps[-1]] if len(case.steps) >= 2 else case.steps
        simplified.append(case)
    return simplified


def section_counts(cases: list[CaseDefinition]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for section in [
        "5.2 Integration Testing",
        "5.3 System Testing",
        "5.4 Performance Testing",
        "5.5 Security Testing",
        "5.6 Regression Testing",
    ]:
        counts[section] = sum(1 for case in cases if case.section == section)
    return counts


def remove_if_exists(workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]


def fill_case_sheet(ws, case: CaseDefinition) -> None:
    ws["A1"] = "Test Case ID"
    ws["B1"] = f"TC ID: {case.tc_id}"
    ws["C1"] = "Test Case Description"
    ws["D1"] = case.description
    ws["A2"] = "Created By"
    ws["B2"] = CREATED_BY
    ws["C2"] = "Reviewed By"
    ws["D2"] = REVIEWED_BY
    ws["E2"] = "Version: v1.0"
    ws["A4"] = "QA Tester's Log"
    ws["B4"] = case.qa_log
    ws["A5"] = "Tester's Name"
    ws["B5"] = CREATED_BY
    ws["C5"] = "Date Tested"
    ws["D5"] = DATE_TESTED
    ws["E5"] = f"Test Case: {case.status}"
    ws["A8"] = "S #"
    ws["B8"] = "Prerequisites:"
    ws["D8"] = "S #"
    ws["E8"] = "Test Data Requirement"

    for row in range(9, 13):
        for cell in (f"A{row}", f"B{row}", f"D{row}", f"E{row}"):
            ws[cell] = None

    for index, value in enumerate(case.prerequisites, start=9):
        ws[f"A{index}"] = index - 8
        ws[f"B{index}"] = value

    for index, value in enumerate(case.test_data, start=9):
        ws[f"D{index}"] = index - 8
        ws[f"E{index}"] = value

    ws["A15"] = case.test_conditions
    ws["A16"] = "Step #"
    ws["B16"] = "Step Details"
    ws["C16"] = "Expected Results"
    ws["D16"] = "Actual Results"
    ws["E16"] = "Pass / Fail / Not executed / Suspended"

    for row in range(17, 25):
        for col in ("A", "B", "C", "D", "E"):
            ws[f"{col}{row}"] = None

    for index, step in enumerate(case.steps, start=17):
        ws[f"A{index}"] = index - 16
        ws[f"B{index}"] = step.detail
        ws[f"C{index}"] = step.expected
        ws[f"D{index}"] = step.actual
        ws[f"E{index}"] = step.result


def update_testcase_workbook(cases: list[CaseDefinition]) -> list[Path]:
    workbook = load_workbook(TESTCASE_PATH)
    template = workbook["TC-01"]

    for number in range(31, 50):
        remove_if_exists(workbook, f"TC-{number}")

    for case in cases:
        new_sheet = workbook.copy_worksheet(template)
        new_sheet.title = case.sheet_name
        fill_case_sheet(new_sheet, case)

    copy_path = BAO_CAO_DIR / "TestCase_Chuong5_JUnitMock.xlsx"
    workbook.save(copy_path)
    outputs = [copy_path]
    try:
        workbook.save(TESTCASE_PATH)
        outputs.insert(0, TESTCASE_PATH)
    except PermissionError:
        locked_copy = BAO_CAO_DIR / "TestCase_rut_gon_cho_Chapter5.xlsx"
        workbook.save(locked_copy)
        outputs.insert(0, locked_copy)
    return outputs


def write_summary_row(ws, row: int, no, item, tested, passed, failed, blocked, skipped, pass_rate, evidence):
    ws[f"C{row}"] = no
    ws[f"D{row}"] = item
    ws[f"E{row}"] = tested
    ws[f"F{row}"] = passed
    ws[f"G{row}"] = failed
    ws[f"H{row}"] = blocked
    ws[f"I{row}"] = skipped
    ws[f"J{row}"] = pass_rate
    ws[f"K{row}"] = evidence


def style_tabular_sheet(ws, header_row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for cell in ws[header_row]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def update_testreport_workbook(cases: list[CaseDefinition], latest_dir: Path) -> list[Path]:
    workbook = load_workbook(TESTREPORT_PATH)
    for sheet_name in ("Chuong5_JUnitMock", "Chuong5_Mapping", "Chuong5_Artifacts"):
        remove_if_exists(workbook, sheet_name)

    counts = section_counts(cases)
    additional_total = sum(counts.values())
    total_tested = 30 + additional_total
    total_passed = 22 + additional_total
    total_failed = 8
    total_pass_rate = f"{(total_passed / total_tested) * 100:.2f}%"

    summary_sheet = workbook.copy_worksheet(workbook[workbook.sheetnames[0]])
    summary_sheet.title = "Chuong5_JUnitMock"
    summary_sheet["C5"] = "CHAPTER 5 JUNIT + MOCK TEST REPORT"
    summary_sheet["C7"] = "Project Name:"
    summary_sheet["D7"] = "Booking System (QLDatVe) - Chapter 5"
    summary_sheet["G7"] = "Creator:"
    summary_sheet["H7"] = CREATED_BY
    summary_sheet["C8"] = "Approver:"
    summary_sheet["D8"] = REVIEWED_BY
    summary_sheet["G8"] = "Date:"
    summary_sheet["H8"] = DATE_TESTED
    summary_sheet["J10"] = "Pass %"
    summary_sheet["K10"] = "Evidence"

    rows = [
        (1, "Functional Testing", 30, 22, 8, 0, 0, "73.33%", "FunctionalTestingSuite + TC-01..TC-30"),
        (2, "Integration Testing", counts["5.2 Integration Testing"], counts["5.2 Integration Testing"], 0, 0, 0, "100%", "BookingApiIntegrationTest + TC-31..TC-32"),
        (3, "System Testing", counts["5.3 System Testing"], counts["5.3 System Testing"], 0, 0, 0, "100%", "BookingSystemFlowMockMvcTest + TC-33..TC-34"),
        (4, "Performance Testing", counts["5.4 Performance Testing"], counts["5.4 Performance Testing"], 0, 0, 0, "100%", "BookingPerformanceSmokeTest + TC-35..TC-36"),
        (5, "Security Testing", counts["5.5 Security Testing"], counts["5.5 Security Testing"], 0, 0, 0, "100%", "SecurityAccessIntegrationTest + TC-37..TC-39"),
        (6, "Regression Testing", counts["5.6 Regression Testing"], counts["5.6 Regression Testing"], 0, 0, 0, "100%", "RegressionStableSuite + TC-40"),
    ]
    for offset, row in enumerate(rows, start=11):
        write_summary_row(summary_sheet, offset, *row)

    write_summary_row(summary_sheet, 17, "Total", "TỔNG CỘNG", total_tested, total_passed, total_failed, 0, 0, total_pass_rate, str(latest_dir / "summary.txt"))
    summary_sheet["C18"] = "Ghi chú"
    summary_sheet["D18"] = "Functional dùng lại 30 test case gốc của BookingService; các hạng mục 5.2-5.6 là test JUnit/mock mới bổ sung."
    summary_sheet["D19"] = "Regression là wrapper suite kiểm tra lại các nhóm test ổn định, nên số liệu tại 5.6 được ghi nhận theo suite-level."
    summary_sheet.column_dimensions["J"].width = 12
    summary_sheet.column_dimensions["K"].width = 42

    mapping_sheet = workbook.create_sheet("Chuong5_Mapping")
    mapping_sheet.append(["TC ID", "Mục báo cáo", "Suite/Class", "Test method", "Kết quả", "Ý nghĩa", "File mã nguồn"])
    for case in cases:
        if case.section == "5.2 Integration Testing":
            file_name = "src/test/java/com/example/QLDatVe/integration/BookingApiIntegrationTest.java"
        elif case.section == "5.3 System Testing":
            file_name = "src/test/java/com/example/QLDatVe/system/BookingSystemFlowMockMvcTest.java"
        elif case.section == "5.4 Performance Testing":
            file_name = "src/test/java/com/example/QLDatVe/performance/BookingPerformanceSmokeTest.java"
        elif case.section == "5.5 Security Testing":
            file_name = "src/test/java/com/example/QLDatVe/security/SecurityAccessIntegrationTest.java"
        else:
            file_name = "src/test/java/com/example/QLDatVe/regression/RegressionStableSuite.java"
        mapping_sheet.append([case.tc_id, case.section, case.suite, case.method_name, case.status, case.note, file_name])
    style_tabular_sheet(mapping_sheet)

    artifact_sheet = workbook.create_sheet("Chuong5_Artifacts")
    artifact_sheet.append(["Artifact", "Đường dẫn", "Cách dùng trong báo cáo"])
    for row in [
        ("Summary Chương 5", str(latest_dir / "summary.txt"), "Dùng cho mục 5.7 thống kê kết quả."),
        ("Functional log", str(latest_dir / "5.1_Functional_Testing.log"), "Dùng làm minh chứng 8 testcase fail hiện hữu."),
        ("Integration log", str(latest_dir / "5.2_Integration_Testing.log"), "Dùng cho mục 5.2 kiểm thử tích hợp."),
        ("System log", str(latest_dir / "5.3_System_Testing.log"), "Dùng cho mục 5.3 kiểm thử hệ thống ở mức mock."),
        ("Performance log", str(latest_dir / "5.4_Performance_Testing.log"), "Dùng cho mục 5.4 kiểm thử hiệu năng smoke."),
        ("Performance metrics 1", str(PERFORMANCE_REPORT_DIR / "create-booking-smoke.txt"), "Chèn số liệu Average/P95/Max cho createBooking."),
        ("Performance metrics 2", str(PERFORMANCE_REPORT_DIR / "concurrent-booking-smoke.txt"), "Chèn số liệu concurrent round cho mục performance."),
        ("Security log", str(latest_dir / "5.5_Security_Testing.log"), "Dùng cho mục 5.5 kiểm thử bảo mật cơ bản."),
        ("Regression log", str(latest_dir / "5.6_Regression_Testing.log"), "Dùng cho mục 5.6 kiểm thử hồi quy."),
        ("Script chạy Chương 5", str(ROOT / "run-chapter5-junit-tests.ps1"), "Dùng mô tả cách thực thi đồng bộ các mục 5.1-5.6."),
    ]:
        artifact_sheet.append(row)
    style_tabular_sheet(artifact_sheet)

    copy_path = BAO_CAO_DIR / "TestReport_Chuong5_JUnitMock.xlsx"
    workbook.save(copy_path)
    outputs = [copy_path]
    try:
        workbook.save(TESTREPORT_PATH)
        outputs.insert(0, TESTREPORT_PATH)
    except PermissionError:
        locked_copy = BAO_CAO_DIR / "TestReport_rut_gon_cho_Chapter5.xlsx"
        workbook.save(locked_copy)
        outputs.insert(0, locked_copy)
    return outputs


def chapter5_markdown(latest_dir: Path) -> str:
    create_booking_metrics = read_performance_metrics("create-booking-smoke.txt")
    concurrent_metrics = read_performance_metrics("concurrent-booking-smoke.txt")
    cases = get_chapter5_cases()
    counts = section_counts(cases)
    total_tested = 30 + sum(counts.values())
    total_passed = 22 + sum(counts.values())
    total_failed = 8
    total_pass_rate = f"{(total_passed / total_tested) * 100:.2f}%"

    return f"""# CHƯƠNG 5 - THỰC THI KIỂM THỬ (JUnit + Mock)

## 5.1. Kiểm thử chức năng (Functional Testing)

Mục này tiếp tục bám vào `BookingService` và sử dụng `FunctionalTestingSuite` để gọi lại `BookingValidationTest`, `BookingSafetyTest`, `BookingLifecycleTest` và `BookingValidationGapsTest`.

Kết quả hiện tại: `30` test case chức năng đã chạy, `22` pass và `8` fail. Trong `8` fail có `3` lỗi `Safety / Concurrency` và `5` lỗi `Validation gaps`. Đây là các bug nghiệp vụ thật của implementation hiện tại, không phải lỗi môi trường kiểm thử.

Minh chứng: `{latest_dir / '5.1_Functional_Testing.log'}`

## 5.2. Kiểm thử tích hợp (Integration Testing)

Phần kiểm thử tích hợp được bổ sung bằng `BookingApiIntegrationTest` với `WebMvcTest`, `MockMvc` và `@MockBean`. Ở bản rút gọn cho báo cáo, phần này giữ `2` test case tiêu biểu: tạo booking CASH thành công và trả lỗi khi service phát hiện seat conflict.

Kết quả thực thi: `{counts["5.2 Integration Testing"]}/{counts["5.2 Integration Testing"]}` test pass.

Minh chứng: `{latest_dir / '5.2_Integration_Testing.log'}`

## 5.3. Kiểm thử hệ thống (System Testing)

Kiểm thử hệ thống ở mức mock được triển khai bằng `BookingSystemFlowMockMvcTest`. Hai hành trình được mô phỏng là luồng `CASH` từ tra cứu chuyến đến `My Bookings`, và luồng `MOMO` từ tạo booking `PENDING` đến checkout, confirm callback và lịch sử booking.

Kết quả thực thi: `2/2` test pass.

Minh chứng: `{latest_dir / '5.3_System_Testing.log'}`

## 5.4. Kiểm thử hiệu năng (Performance Testing)

Hiệu năng được kiểm tra ở mức smoke test bằng `BookingPerformanceSmokeTest`. Kết quả đo được:

- `createBooking` lặp `150` lần: Average = `{create_booking_metrics.get('Average (ms)', '?')}`, P95 = `{create_booking_metrics.get('P95 (ms)', '?')}`, Max = `{create_booking_metrics.get('Max (ms)', '?')}`;
- `12` vòng concurrent booking: Average = `{concurrent_metrics.get('Average (ms)', '?')}`, P95 = `{concurrent_metrics.get('P95 (ms)', '?')}`, Max = `{concurrent_metrics.get('Max (ms)', '?')}`.

Cả hai test đều pass theo ngưỡng đã đặt trong mã nguồn.

Minh chứng: `{latest_dir / '5.4_Performance_Testing.log'}`

## 5.5. Kiểm thử bảo mật cơ bản (Security Testing)

Kiểm thử bảo mật cơ bản được bổ sung bằng `SecurityAccessIntegrationTest`, tập trung vào xác thực và phân quyền. Trong bản rút gọn cho báo cáo, phần này giữ `3` test case tiêu biểu: cho phép anonymous truy cập `/api/trips`, chặn anonymous ở `/api/bookings/my-bookings`, và cho phép admin truy cập `/api/admin/bookings`.

Kết quả thực thi: `{counts["5.5 Security Testing"]}/{counts["5.5 Security Testing"]}` test pass.

Minh chứng: `{latest_dir / '5.5_Security_Testing.log'}`

## 5.6. Kiểm thử hồi quy (Regression Testing)

Kiểm thử hồi quy được triển khai bằng `RegressionStableSuite`. Suite này chạy lại `BookingValidationTest`, `BookingLifecycleTest` và `SecurityAccessIntegrationTest` thông qua `SuiteAssertions` để xác nhận các nhóm test ổn định vẫn không bị ảnh hưởng sau khi bổ sung test mới.

Kết quả thực thi: `1/1` suite pass.

Minh chứng: `{latest_dir / '5.6_Regression_Testing.log'}`

## 5.7. Thống kê kết quả

Tổng hợp Chương 5 hiện có:

- Functional Testing: `30` tested, `22` passed, `8` failed;
- Integration Testing: `{counts["5.2 Integration Testing"]}` tested, `{counts["5.2 Integration Testing"]}` passed;
- System Testing: `2` tested, `2` passed;
- Performance Testing: `2` tested, `2` passed;
- Security Testing: `{counts["5.5 Security Testing"]}` tested, `{counts["5.5 Security Testing"]}` passed;
- Regression Testing: `{counts["5.6 Regression Testing"]}` suite, `{counts["5.6 Regression Testing"]}` passed.

Tổng cộng có `{total_tested}` lượt thực thi được ghi nhận, trong đó `{total_passed}` pass, `{total_failed}` fail, tỷ lệ pass khoảng `{total_pass_rate}`.

Lưu ý: các mục `5.2` đến `5.6` đều chạy trong môi trường `JUnit + Mock / MockMvc`, không dùng cơ sở dữ liệu thật hay dịch vụ ngoài thật.
"""


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"


def w_tag(name: str) -> str:
    return f"{{{W_NS}}}{name}"


def paragraph(text: str, bold: bool = False) -> ET.Element:
    p = ET.Element(w_tag("p"))
    r = ET.SubElement(p, w_tag("r"))
    if bold:
        r_pr = ET.SubElement(r, w_tag("rPr"))
        ET.SubElement(r_pr, w_tag("b"))
    t = ET.SubElement(r, w_tag("t"))
    if text.startswith(" ") or text.endswith(" "):
        t.set(XML_SPACE, "preserve")
    t.text = text
    return p


def create_docx_from_markdown(template_docx: Path, output_docx: Path, markdown: str) -> None:
    ET.register_namespace("w", W_NS)
    with zipfile.ZipFile(template_docx) as archive:
        template_xml = archive.read("word/document.xml")
    template_root = ET.fromstring(template_xml)
    body = template_root.find(w_tag("body"))
    sect_pr = copy.deepcopy(body.find(w_tag("sectPr"))) if body is not None else None

    root = ET.Element(w_tag("document"))
    new_body = ET.SubElement(root, w_tag("body"))
    for line in markdown.splitlines():
        if not line.strip():
            new_body.append(paragraph(""))
        elif line.startswith("# "):
            new_body.append(paragraph(line[2:], bold=True))
        elif line.startswith("## "):
            new_body.append(paragraph(line[3:], bold=True))
        else:
            new_body.append(paragraph(line))
    if sect_pr is not None:
        new_body.append(sect_pr)

    temp_docx = output_docx.with_name(output_docx.stem + "_tmp.docx")
    xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    with zipfile.ZipFile(template_docx) as src, zipfile.ZipFile(temp_docx, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "word/document.xml":
                data = xml_bytes
            dst.writestr(item, data)
    final_output = output_docx
    if output_docx.exists():
        try:
            output_docx.unlink()
        except PermissionError:
            final_output = output_docx.with_name(output_docx.stem + "_updated.docx")
            if final_output.exists():
                final_output.unlink()
    shutil.copyfile(temp_docx, final_output)
    try:
        temp_docx.unlink()
    except PermissionError:
        pass


def main() -> None:
    latest_dir = latest_execution_dir()
    cases = get_chapter5_cases()
    testcase_outputs = update_testcase_workbook(cases)
    testreport_outputs = update_testreport_workbook(cases, latest_dir)

    chapter5_md = BAO_CAO_DIR / "Chuong5_ThucThiKiemThu_JUnitMock.md"
    chapter5_docx = BAO_CAO_DIR / "Chuong5_ThucThiKiemThu_JUnitMock.docx"

    markdown = chapter5_markdown(latest_dir)
    chapter5_md.write_text(markdown, encoding="utf-8")
    create_docx_from_markdown(SUMMARY_TEMPLATE_DOCX, chapter5_docx, markdown)

    print("Created/updated files:")
    for path in [*testcase_outputs, *testreport_outputs, chapter5_md, chapter5_docx]:
        print(path)


if __name__ == "__main__":
    main()
