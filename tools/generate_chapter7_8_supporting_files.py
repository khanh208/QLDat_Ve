from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path

from docx import Document
from docx.enum.table import WD_TABLE_ALIGNMENT
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.shared import Pt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
BAO_CAO_DIR = ROOT / "Bao_Cao"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


@dataclass(frozen=True)
class BugRow:
    bug_id: str
    group: str
    title: str
    severity: str
    priority: str
    evidence: str
    status: str


BUGS: list[BugRow] = [
    BugRow(
        "BUG-01",
        "Safety / Concurrency",
        "Hai người dùng có thể cùng đặt thành công một ghế bằng CASH trong tình huống đồng thời.",
        "Critical",
        "P1",
        "BookingSafetyTest.testSafe_RaceCondition_CASH_OneSeat",
        "Open",
    ),
    BugRow(
        "BUG-02",
        "Safety / Concurrency",
        "Hai người dùng có thể cùng đặt thành công một ghế bằng MOMO trong tình huống đồng thời.",
        "Critical",
        "P1",
        "BookingSafetyTest.testSafe_RaceCondition_MOMO_OneSeat",
        "Open",
    ),
    BugRow(
        "BUG-03",
        "Safety / Concurrency",
        "Hai request có tập ghế giao nhau vẫn có thể cùng được lưu, gây trùng ghế.",
        "Critical",
        "P1",
        "BookingSafetyTest.testSafe_RaceCondition_OverlapSeats",
        "Open",
    ),
    BugRow(
        "BUG-04",
        "Validation",
        "paymentMethod = null chưa bị từ chối đúng theo yêu cầu nghiệp vụ.",
        "High",
        "P2",
        "BookingValidationGapsTest.testValGap_PaymentMethodNull_ShouldReject",
        "Open",
    ),
    BugRow(
        "BUG-05",
        "Validation",
        "paymentMethod không hợp lệ như BANKING vẫn có thể đi qua service.",
        "High",
        "P2",
        "BookingValidationGapsTest.testValGap_PaymentMethodUnsupported_ShouldReject",
        "Open",
    ),
    BugRow(
        "BUG-06",
        "Validation",
        "seatNumbers chứa giá trị null chưa bị chặn ở lớp service.",
        "Medium",
        "P2",
        "BookingValidationGapsTest.testValGap_SeatContainsNull_ShouldReject",
        "Open",
    ),
    BugRow(
        "BUG-07",
        "Validation",
        "seatNumbers chứa chuỗi rỗng chưa bị từ chối đúng cách.",
        "Medium",
        "P2",
        "BookingValidationGapsTest.testValGap_SeatContainsBlank_ShouldReject",
        "Open",
    ),
    BugRow(
        "BUG-08",
        "Validation",
        "Request có cả ghế hợp lệ và chuỗi rỗng vẫn chưa bị chặn hoàn toàn.",
        "Medium",
        "P2",
        "BookingValidationGapsTest.testValGap_SeatContainsValidAndBlank_ShouldReject",
        "Open",
    ),
]


CHAPTER7_TEXT: list[tuple[str, str]] = [
    ("CHƯƠNG 7: BÁO CÁO LỖI (BUG REPORT)", "Heading 1"),
    ("7.1. Mục đích của báo cáo lỗi", "Heading 2"),
    (
        "Mục đích của chương này là tổng hợp các lỗi đã được phát hiện trong quá trình thực thi kiểm thử đối với module BookingService và các flow liên quan. Phần báo cáo lỗi không chỉ liệt kê các testcase fail mà còn giúp chỉ ra nhóm lỗi nào có mức độ ảnh hưởng nghiêm trọng, lỗi nào cần ưu tiên sửa trước và bằng chứng nào đang được dùng để khẳng định sự tồn tại của lỗi.",
        "Normal",
    ),
    (
        "Trong phạm vi dự án hiện tại, các lỗi được ghi nhận chủ yếu xuất phát từ bộ kiểm thử chức năng, đặc biệt là các testcase về Safety, Concurrency và Validation. Đây đều là những lỗi có liên quan trực tiếp đến độ tin cậy của nghiệp vụ đặt vé, vì vậy việc phân tích riêng chúng trong một chương độc lập là cần thiết.",
        "Normal",
    ),
    ("7.2. Danh sách lỗi đã phát hiện", "Heading 2"),
    (
        "Từ kết quả thực thi bộ FunctionalTestingSuite, hiện có tổng cộng 8 lỗi đang được ghi nhận. Trong đó có 3 lỗi thuộc nhóm Safety / Concurrency và 5 lỗi thuộc nhóm Validation đầu vào. Các lỗi này đều đã có testcase tái hiện cụ thể trong mã nguồn và được phản ánh trực tiếp qua log thực thi tại thư mục target/chapter5-test-execution cùng các báo cáo surefire.",
        "Normal",
    ),
    ("7.2.1. Nhóm lỗi Safety / Concurrency", "Heading 2"),
    (
        "Nhóm lỗi Safety / Concurrency gồm ba lỗi nghiêm trọng nhất của hệ thống hiện tại. Lỗi thứ nhất xảy ra khi hai người dùng cùng đặt một ghế bằng phương thức CASH tại cùng một thời điểm, dẫn đến khả năng cả hai booking đều được lưu. Lỗi thứ hai có bản chất tương tự nhưng xảy ra với phương thức MOMO. Lỗi thứ ba xảy ra khi hai request đặt các tập ghế giao nhau, trong đó có ít nhất một ghế bị trùng, nhưng hệ thống vẫn có thể cho phép cả hai booking thành công. Ba lỗi này đều cho thấy luồng kiểm tra ghế và lưu booking chưa được thực hiện theo cách đủ an toàn trong môi trường đồng thời.",
        "Normal",
    ),
    (
        "Các lỗi concurrency trên được xem là mức độ Critical vì chúng tác động trực tiếp đến thuộc tính Safety của bài toán đặt vé, tức hệ thống có nguy cơ bán trùng ghế hoặc bán vượt số ghế. Nếu lỗi này xảy ra trong môi trường thực tế, hậu quả không chỉ là dữ liệu không nhất quán mà còn ảnh hưởng trực tiếp đến người dùng cuối và uy tín của hệ thống.",
        "Normal",
    ),
    ("7.2.2. Nhóm lỗi Validation đầu vào", "Heading 2"),
    (
        "Nhóm lỗi Validation hiện có năm trường hợp chính. Cụ thể, service hiện chưa chặn đúng các request khi paymentMethod bằng null, paymentMethod không thuộc tập giá trị hợp lệ, danh sách ghế chứa null, danh sách ghế chứa chuỗi rỗng, hoặc danh sách ghế chứa đồng thời giá trị hợp lệ và chuỗi rỗng. Các testcase này được thiết kế với mong đợi hệ thống phải từ chối request, tuy nhiên implementation hiện tại vẫn để request đi qua và tạo booking trong một số tình huống.",
        "Normal",
    ),
    (
        "So với lỗi concurrency, nhóm lỗi validation có mức độ ảnh hưởng thấp hơn nhưng vẫn cần được xử lý sớm. Nếu các request đầu vào không hợp lệ tiếp tục được chấp nhận, hệ thống có thể phát sinh dữ liệu bẩn, làm tăng độ khó khi bảo trì và làm suy giảm độ tin cậy của các bước xử lý phía sau như thanh toán, xác nhận và thống kê booking.",
        "Normal",
    ),
    ("7.3. Mức độ ảnh hưởng và ưu tiên xử lý", "Heading 2"),
    (
        "Thứ tự ưu tiên xử lý nên bắt đầu từ ba lỗi Safety / Concurrency vì đây là các lỗi P1 và ảnh hưởng trực tiếp đến nghiệp vụ cốt lõi của hệ thống đặt vé. Sau đó, năm lỗi Validation nên được xử lý tiếp theo để làm chặt lớp kiểm tra đầu vào và giảm nguy cơ tạo booking không hợp lệ. Nếu phải lựa chọn theo thứ tự, hệ thống nên ưu tiên khóa được lỗi đặt trùng ghế trước, rồi mới hoàn thiện dần các ràng buộc validation.",
        "Normal",
    ),
    ("7.4. Minh chứng và cách tái hiện lỗi", "Heading 2"),
    (
        "Minh chứng chính cho các lỗi ở chương này đến từ các testcase đã được cài đặt trong BookingSafetyTest và BookingValidationGapsTest. Ngoài mã nguồn test, người thực hiện còn có thể đối chiếu với log trong thư mục target/chapter5-test-execution, file summary.txt của lần chạy gần nhất và các báo cáo chi tiết trong target/surefire-reports. Với nhóm lỗi concurrency ở frontend, workflow selenium-known-issues.yml còn có thể được dùng để sinh ảnh chụp màn hình và evidence bundle cho các testcase TC02, TC03 và TC24 nhằm hỗ trợ minh chứng trực quan hơn.",
        "Normal",
    ),
]


CHAPTER8_TEXT: list[tuple[str, str]] = [
    ("CHƯƠNG 8: TỔNG KẾT VÀ ĐÁNH GIÁ", "Heading 1"),
    ("8.1. Kết quả đạt được", "Heading 2"),
    (
        "Qua quá trình thực hiện, báo cáo đã xây dựng được một bộ kiểm thử tương đối đầy đủ cho bài toán đặt vé tập trung vào module BookingService. Ở mức nghiệp vụ, bộ test hiện tại đã bao phủ các luồng đặt vé thành công, các trường hợp dữ liệu không hợp lệ, các nhánh vòng đời booking và các tình huống đồng thời có nguy cơ gây bán trùng ghế. Ở mức công cụ, dự án đã kết hợp được JUnit, Mockito, MockMvc, Selenium, GitHub Actions và đặc tả TLA+ để mở rộng góc nhìn đánh giá chất lượng hệ thống.",
        "Normal",
    ),
    (
        "Về kết quả thực thi, bộ kiểm thử chức năng trọng tâm cho BookingService hiện ghi nhận 30 testcase với 22 testcase pass và 8 testcase fail. Ngoài ra, phần kiểm thử bổ sung bằng JUnit + Mock trong Chương 5 ghi nhận tổng 40 lượt thực thi với 32 pass và 8 fail. Phần automation trong Chương 6 cũng đã được cấu hình để chạy lặp lại trên GitHub Actions, đặc biệt là UI smoke test theo matrix Chrome và Firefox.",
        "Normal",
    ),
    ("8.2. Đánh giá chất lượng hiện tại của hệ thống", "Heading 2"),
    (
        "Chất lượng hiện tại của hệ thống có thể được đánh giá ở mức tương đối tốt đối với các luồng nghiệp vụ cơ bản như tạo booking hợp lệ, xử lý vòng đời booking, xác nhận thanh toán, hủy booking và dọn dẹp các booking PENDING quá hạn. Các kết quả pass ở nhóm lifecycle, integration, system, performance smoke, security cơ bản và regression cho thấy phần lớn logic nền tảng của hệ thống đã hoạt động ổn định trong phạm vi kiểm thử hiện có.",
        "Normal",
    ),
    (
        "Tuy nhiên, hệ thống vẫn chưa đạt yêu cầu ở thuộc tính Safety vì các testcase đồng thời đã chỉ ra nguy cơ bán trùng ghế khi nhiều request xảy ra cùng lúc. Bên cạnh đó, năm lỗi validation đầu vào cho thấy lớp service hiện vẫn còn thiếu một số ràng buộc cần thiết. Vì vậy, có thể kết luận rằng hệ thống đã tốt ở nhiều luồng chuẩn nhưng vẫn còn một số rủi ro quan trọng cần xử lý trước khi có thể xem là đủ an toàn cho môi trường thực tế.",
        "Normal",
    ),
    ("8.3. Hạn chế của báo cáo và phạm vi hiện tại", "Heading 2"),
    (
        "Báo cáo này vẫn có một số giới hạn cần ghi rõ. Thứ nhất, trọng tâm kiểm thử vẫn tập trung chủ yếu vào module BookingService ở backend, chưa mở rộng thành một đợt end-to-end test đầy đủ với cơ sở dữ liệu thật, cổng thanh toán thật và email thật. Thứ hai, một phần automation frontend hiện đang dựa trên mock API để kiểm tra luồng giao diện, do đó giá trị của nó là rất tốt cho regression và minh chứng UI, nhưng chưa thay thế hoàn toàn cho môi trường tích hợp thật.",
        "Normal",
    ),
    (
        "Ngoài ra, dù đề tài có định hướng TLA+ và phần đặc tả đã được bổ sung vào repo, báo cáo hiện vẫn chủ yếu dựa trên bằng chứng kiểm thử thực nghiệm và chưa trình bày kết quả model checking TLC đầy đủ. Đây là điểm cần nêu rõ để bảo đảm tính trung thực và nhất quán giữa phần mô tả đề tài với hiện vật thực tế có trong dự án.",
        "Normal",
    ),
    ("8.4. Hướng cải tiến trong tương lai", "Heading 2"),
    (
        "Trong giai đoạn tiếp theo, hệ thống nên ưu tiên xử lý triệt để các lỗi Safety bằng cách tăng tính nguyên tử cho luồng kiểm tra ghế và lưu booking, kết hợp với các biện pháp như locking phù hợp hoặc ràng buộc dữ liệu ở mức database. Song song với đó, lớp validation cần được siết chặt để từ chối sớm các request có paymentMethod hoặc seatNumbers không hợp lệ.",
        "Normal",
    ),
    (
        "Về mặt kiểm thử, nên tiếp tục mở rộng regression UI, tăng số lượng integration test với môi trường gần thật hơn, và nếu điều kiện cho phép thì bổ sung kết quả model checking TLC cho phần TLA+ để củng cố chiều sâu học thuật của đề tài. Đây sẽ là các bước giúp bài báo cáo vừa bám sát dự án thực tế, vừa thể hiện rõ giá trị của kiểm thử phần mềm trong việc phát hiện và kiểm soát rủi ro hệ thống.",
        "Normal",
    ),
]


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def get_bao_cao_path(name: str) -> Path:
    direct = BAO_CAO_DIR / name
    if direct.exists():
        return direct
    for path in BAO_CAO_DIR.iterdir():
        if path.name == name:
            return path
    return direct


def create_bug_workbook() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "BugSummary"
    ws.append(["Bug ID", "Nhóm lỗi", "Mô tả", "Severity", "Priority", "Evidence", "Status"])
    for bug in BUGS:
        ws.append([bug.bug_id, bug.group, bug.title, bug.severity, bug.priority, bug.evidence, bug.status])
    style_sheet(ws)

    ws_group = wb.create_sheet("ByGroup")
    ws_group.append(["Nhóm lỗi", "Số lượng", "Ghi chú"])
    ws_group.append(["Safety / Concurrency", 3, "Nhóm lỗi nghiêm trọng nhất, ảnh hưởng trực tiếp đến thuộc tính Safety."])
    ws_group.append(["Validation", 5, "Nhóm lỗi đầu vào chưa được chặn chặt ở lớp service."])
    style_sheet(ws_group)

    output = BAO_CAO_DIR / "Bảng thống kê Chương 7 Bug Report.xlsx"
    wb.save(output)
    return output


def update_testreport_workbook() -> Path:
    report_path = get_bao_cao_path("TestReport.xlsx")
    workbook = load_workbook(report_path)

    for sheet_name in ["Chuong7_BugSummary", "Chuong8_TongKet"]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    ws = workbook.create_sheet("Chuong7_BugSummary")
    ws.append(["Bug ID", "Nhóm lỗi", "Mô tả", "Severity", "Priority", "Evidence", "Status"])
    for bug in BUGS:
        ws.append([bug.bug_id, bug.group, bug.title, bug.severity, bug.priority, bug.evidence, bug.status])
    style_sheet(ws)

    ws2 = workbook.create_sheet("Chuong8_TongKet")
    ws2.append(["Chỉ số", "Giá trị", "Ghi chú"])
    rows = [
        ("Functional core tests", "30 executed | 22 pass | 8 fail", "Bộ kiểm thử trọng tâm cho BookingService."),
        ("Chapter 5 JUnit + Mock", "40 executed | 32 pass | 8 fail", "Bao gồm functional, integration, system, performance, security, regression."),
        ("Nhóm bug nghiêm trọng", "3 lỗi Safety / Concurrency", "Cần ưu tiên xử lý."),
        ("Nhóm bug validation", "5 lỗi validation", "Cần hoàn thiện kiểm tra đầu vào."),
        ("Automation CI", "4 workflow", "backend-regression, performance-smoke, selenium-e2e, selenium-known-issues."),
    ]
    for row in rows:
        ws2.append(list(row))
    style_sheet(ws2)

    try:
        workbook.save(report_path)
        return report_path
    except PermissionError:
        fallback = BAO_CAO_DIR / "TestReport_Chuong7_8.xlsx"
        workbook.save(fallback)
        return fallback


def create_docx(path: Path, title: str, sections: list[tuple[str, str]]) -> Path:
    doc = Document()
    for text, style in sections:
        para = doc.add_paragraph(text)
        para.style = style
        if style == "Heading 1":
            para.alignment = WD_ALIGN_PARAGRAPH.CENTER
            for run in para.runs:
                run.font.size = Pt(16)
                run.bold = True
    doc.save(path)
    return path


def update_markdown_inventory(files: list[Path]) -> Path:
    path = BAO_CAO_DIR / "du-lieu-bam-sat-du-an-cho-bao-cao-word.md"
    marker = "\n## Hiện vật Chương 7 và Chương 8\n"
    text = path.read_text(encoding="utf-8")
    block = marker + "\n".join(f"- `{file.name}`" for file in files) + "\n"
    if marker in text:
        text = text.split(marker)[0].rstrip() + "\n" + block
    else:
        text += "\n" + block
    path.write_text(text, encoding="utf-8")
    return path


def main() -> None:
    outputs = [
        create_bug_workbook(),
        update_testreport_workbook(),
        create_docx(BAO_CAO_DIR / "Chuong7_BaoCaoLoi.docx", "Chuong7", CHAPTER7_TEXT),
        create_docx(BAO_CAO_DIR / "Chuong8_TongKetVaDanhGia.docx", "Chuong8", CHAPTER8_TEXT),
    ]
    inventory = update_markdown_inventory(outputs)
    outputs.append(inventory)

    print("Created Chapter 7/8 supporting files:")
    for output in outputs:
        print(output)


if __name__ == "__main__":
    main()
