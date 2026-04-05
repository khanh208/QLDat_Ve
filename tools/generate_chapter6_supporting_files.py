from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path

from docx import Document
from docx.enum.table import WD_TABLE_ALIGNMENT
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.shared import Pt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
BAO_CAO_DIR = ROOT / "Bao_Cao"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


@dataclass(frozen=True)
class WorkflowRow:
    area: str
    workflow: str
    trigger: str
    tools: str
    scope: str
    output: str
    note: str


@dataclass(frozen=True)
class AutomationFlow:
    flow_id: str
    category: str
    test_class: str
    test_method: str
    objective: str
    browsers: str


WORKFLOWS: list[WorkflowRow] = [
    WorkflowRow(
        area="Hồi quy backend ổn định",
        workflow="backend-regression.yml",
        trigger="push, pull_request, workflow_dispatch",
        tools="JUnit 4, Maven Surefire, GitHub Actions",
        scope="RegressionStableSuite (1 suite ổn định)",
        output="target/surefire-reports",
        note="Dùng để phát hiện sớm ảnh hưởng khi code backend thay đổi.",
    ),
    WorkflowRow(
        area="Hiệu năng smoke",
        workflow="performance-smoke.yml",
        trigger="workflow_dispatch",
        tools="JUnit 4, Mockito, GitHub Actions",
        scope="BookingPerformanceSmokeTest (2 test)",
        output="target/performance-reports, target/surefire-reports",
        note="Chạy theo yêu cầu để đo nhanh createBooking và kịch bản đồng thời.",
    ),
    WorkflowRow(
        area="UI smoke test",
        workflow="selenium-e2e.yml",
        trigger="push, workflow_dispatch",
        tools="Selenium WebDriver, Node mock API, React dev server, GitHub Actions matrix",
        scope="FrontendSeleniumSmokeTest (6 flow x 2 browser = 12 lượt mỗi run)",
        output="frontend.log, mock-api.log, target/selenium-screenshots, target/surefire-reports",
        note="Chạy song song trên Chrome và Firefox bằng matrix strategy.",
    ),
    WorkflowRow(
        area="Minh chứng known issues",
        workflow="selenium-known-issues.yml",
        trigger="workflow_dispatch",
        tools="Selenium WebDriver, Node mock API, GitHub Actions matrix",
        scope="FrontendKnownIssueEvidenceTest (3 flow x 2 browser = 6 lượt mỗi run)",
        output="frontend.log, mock-api.log, target/selenium-screenshots, target/surefire-reports",
        note="Dùng continue-on-error để vẫn thu được ảnh/log khi testcase chủ đích fail.",
    ),
]


FLOWS: list[AutomationFlow] = [
    AutomationFlow(
        flow_id="AT-01",
        category="Selenium smoke",
        test_class="FrontendSeleniumSmokeTest",
        test_method="homePageShouldRenderFeaturedTrips",
        objective="Kiểm tra trang chủ hiển thị danh sách chuyến.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-02",
        category="Selenium smoke",
        test_class="FrontendSeleniumSmokeTest",
        test_method="cashBookingFlowShouldShowSuccessNotification",
        objective="Kiểm tra luồng đặt vé CASH thành công.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-03",
        category="Selenium smoke",
        test_class="FrontendSeleniumSmokeTest",
        test_method="bookedSeatShouldBeMarkedUnavailable",
        objective="Kiểm tra ghế đã bán bị khóa trên giao diện.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-04",
        category="Selenium smoke",
        test_class="FrontendSeleniumSmokeTest",
        test_method="unauthenticatedUserShouldBeRedirectedToLoginForMyBookings",
        objective="Kiểm tra người dùng chưa đăng nhập bị chuyển về login.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-05",
        category="Selenium smoke",
        test_class="FrontendSeleniumSmokeTest",
        test_method="myBookingsPageShouldRenderExistingBookingsForLoggedInUser",
        objective="Kiểm tra trang My Bookings hiển thị dữ liệu khi đã đăng nhập.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-06",
        category="Selenium smoke",
        test_class="FrontendSeleniumSmokeTest",
        test_method="momoBookingFlowShouldReachPaymentSuccessPage",
        objective="Kiểm tra luồng MOMO đi tới Payment Success.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-07",
        category="Selenium evidence",
        test_class="FrontendKnownIssueEvidenceTest",
        test_method="tc02_sameSeatMomo_shouldRejectOneBooking",
        objective="Hai user cùng đặt 1 ghế bằng MOMO; nếu cả hai đều thành công thì testcase fail.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-08",
        category="Selenium evidence",
        test_class="FrontendKnownIssueEvidenceTest",
        test_method="tc03_overlapSeatCash_shouldRejectOneBooking",
        objective="Hai user đặt tập ghế giao nhau bằng CASH; nếu cả hai đều thành công thì testcase fail.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-09",
        category="Selenium evidence",
        test_class="FrontendKnownIssueEvidenceTest",
        test_method="tc24_differentSeatsCash_shouldAllowBothBookings",
        objective="Hai user đặt đồng thời nhưng khác ghế bằng CASH; cả hai phải thành công.",
        browsers="chrome, firefox",
    ),
    AutomationFlow(
        flow_id="AT-10",
        category="Backend automation",
        test_class="RegressionStableSuite / BookingPerformanceSmokeTest",
        test_method="suite + smoke",
        objective="Chạy tự động suite hồi quy ổn định và hiệu năng smoke trong CI.",
        browsers="Không áp dụng",
    ),
]


ARTIFACT_ROWS = [
    ("Log frontend", "frontend.log", "Ghi lại log khởi động React frontend trong workflow Selenium."),
    ("Log mock API", "mock-api.log", "Ghi lại log mock server dùng cho Selenium và evidence workflow."),
    ("Surefire reports", "target/surefire-reports", "Lưu kết quả XML/TXT của JUnit test và Selenium test."),
    ("Performance reports", "target/performance-reports", "Lưu số liệu Average, P95 và Max của performance smoke test."),
    ("Selenium screenshots", "target/selenium-screenshots/pass|fail", "Lưu ảnh checkpoint khi pass và ảnh lỗi khi fail theo từng nhóm home, booking, auth, my-bookings, payment."),
    ("Evidence bundle", "target/selenium-screenshots/evidence/TC02|TC03|TC24", "Lưu ảnh before/after và summary cho các testcase known issue."),
]


CHAPTER6_TEXT: list[tuple[str, str]] = [
    ("CHƯƠNG 6: KIỂM THỬ TỰ ĐỘNG (AUTOMATION TESTING)", "Heading 1"),
    ("6.1. Mục tiêu kiểm thử tự động", "Heading 2"),
    (
        "Kiểm thử tự động trong dự án QLDatVe được bổ sung nhằm giảm khối lượng chạy test thủ công lặp lại, giúp nhóm có thể kiểm tra nhanh các nhánh nghiệp vụ quan trọng sau mỗi lần cập nhật mã nguồn. Phần automation cũng đóng vai trò chuẩn hóa cách thực thi test, tạo ra log, báo cáo và ảnh chụp màn hình có cấu trúc để phục vụ phân tích kết quả trong báo cáo.",
        "Normal",
    ),
    (
        "Phạm vi tự động hóa hiện tại tập trung vào ba lớp chính. Thứ nhất là automation backend bằng JUnit, Mockito và MockMvc cho các luồng ổn định của BookingService. Thứ hai là automation hiệu năng ở mức smoke test để đo nhanh thời gian xử lý của createBooking. Thứ ba là automation giao diện bằng Selenium kết hợp GitHub Actions, dùng để kiểm tra các luồng người dùng quan trọng trên frontend như xem chuyến xe, đặt vé CASH, đặt vé MOMO, kiểm tra ghế đã bán và trang lịch sử vé.",
        "Normal",
    ),
    ("6.2. Công cụ và kiến trúc automation", "Heading 2"),
    (
        "Hệ thống automation của dự án được xây dựng trên nền JUnit 4 và Maven Surefire ở phía backend, kết hợp với Mockito và MockMvc để tách logic nghiệp vụ khỏi dependency thật. Đối với frontend, dự án sử dụng Selenium WebDriver để điều khiển trình duyệt, Node.js để chạy mock API server, React dev server để dựng giao diện và GitHub Actions để điều phối toàn bộ pipeline kiểm thử tự động trên môi trường CI.",
        "Normal",
    ),
    (
        "Điểm đáng chú ý của phần automation là workflow Selenium được cấu hình theo matrix strategy để chạy song song trên hai trình duyệt Chrome và Firefox. Cách tổ chức này giúp tăng độ tin cậy của kiểm thử giao diện, giảm nguy cơ testcase chỉ pass ở một browser duy nhất và tạo ra bộ minh chứng có giá trị hơn cho báo cáo. Ngoài workflow smoke test ổn định, dự án còn có workflow riêng để thu thập bằng chứng cho các lỗi known issue như đặt trùng ghế trong tình huống hai người dùng thao tác đồng thời.",
        "Normal",
    ),
    ("6.3. Automation backend bằng JUnit + Mock", "Heading 2"),
    (
        "Ở phía backend, dự án hiện có workflow backend-regression.yml để chạy RegressionStableSuite mỗi khi push code, mở pull request hoặc kích hoạt thủ công. Workflow này dùng để phát hiện nhanh việc thay đổi mã nguồn có làm ảnh hưởng đến các nhánh nghiệp vụ ổn định hay không. Song song với đó, workflow performance-smoke.yml cho phép chạy BookingPerformanceSmokeTest theo yêu cầu để ghi nhận số liệu hiệu năng smoke test và upload báo cáo ra artifact.",
        "Normal",
    ),
    (
        "Ngoài GitHub Actions, dự án còn duy trì khả năng chạy automation backend ngay trên máy local thông qua Maven và script PowerShell. Điều này giúp người thực hiện có thể chủ động chạy lại các suite functional, integration, system, performance, security và regression trước khi commit code, từ đó giảm xác suất đưa lỗi lên môi trường CI.",
        "Normal",
    ),
    ("6.4. Automation UI bằng Selenium + GitHub Actions", "Heading 2"),
    (
        "Workflow selenium-e2e.yml được cấu hình để chạy khi push code và khi kích hoạt thủ công. Trong mỗi lần chạy, workflow lần lượt checkout mã nguồn, cài Java và Node, cài dependency frontend, khởi động mock API server, khởi động React frontend, chờ các dịch vụ sẵn sàng rồi mới chạy bộ FrontendSeleniumSmokeTest. Bộ smoke test hiện tại bao phủ sáu luồng chính gồm hiển thị danh sách chuyến, đặt vé CASH, kiểm tra ghế đã bán, chuyển hướng login, hiển thị My Bookings và luồng thanh toán MOMO.",
        "Normal",
    ),
    (
        "Để phục vụ phân tích lỗi và minh chứng báo cáo, workflow selenium-known-issues.yml được bổ sung riêng cho ba testcase quan trọng là TC02, TC03 và TC24. Đây là các ca kiểm thử mô phỏng hai người dùng thao tác đồng thời trên cùng ghế hoặc trên tập ghế giao nhau. Workflow này sử dụng continue-on-error để trong trường hợp testcase fail theo đúng bug hiện hữu, hệ thống vẫn tiếp tục upload ảnh chụp màn hình, summary và log để người kiểm thử có thể sử dụng làm bằng chứng trong báo cáo lỗi.",
        "Normal",
    ),
    ("6.5. Artifact và đầu ra của automation", "Heading 2"),
    (
        "Mỗi workflow automation đều tạo ra artifact riêng để phục vụ kiểm tra kết quả. Với backend regression, artifact chính là thư mục target/surefire-reports. Với performance smoke, dự án lưu thêm target/performance-reports chứa số liệu Average, P95 và Max. Với Selenium, artifact bao gồm frontend.log, mock-api.log, target/surefire-reports và target/selenium-screenshots.",
        "Normal",
    ),
    (
        "Thư mục target/selenium-screenshots được tổ chức theo pass và fail, sau đó tách tiếp theo từng thành phần như home, booking, auth, my-bookings và payment. Cách tổ chức này giúp người thực hiện có thể nhanh chóng tìm lại ảnh minh họa cho từng testcase. Riêng workflow known issue còn tạo evidence bundle cho từng testcase TC02, TC03 và TC24 để làm rõ expected result, observed result và trạng thái của từng user trước và sau khi submit booking.",
        "Normal",
    ),
    ("6.6. Đánh giá chung", "Heading 2"),
    (
        "Nhìn chung, phần kiểm thử tự động của dự án đã được tổ chức theo hướng thực tế và có thể sử dụng lại nhiều lần. Backend có regression workflow để kiểm tra nhanh các luồng ổn định, performance smoke để quan sát hiệu năng cơ bản, còn frontend có Selenium smoke test chạy trên nhiều trình duyệt ngay trong GitHub Actions. Đây là bước tiến quan trọng so với việc chỉ chạy test thủ công hoặc chỉ chạy unit test rời rạc trên máy cá nhân.",
        "Normal",
    ),
    (
        "Tuy vậy, phần automation hiện vẫn chủ yếu dựa trên mock đối với frontend flow và chưa thay thế hoàn toàn cho end-to-end test với môi trường thật, cơ sở dữ liệu thật và cổng thanh toán thật. Do đó, hướng phát triển tiếp theo là tiếp tục mở rộng phạm vi regression UI, bổ sung thêm kịch bản booking đồng thời ở mức hệ thống thật và tăng cường việc lưu trữ artifact để phục vụ phân tích lỗi sâu hơn.",
        "Normal",
    ),
]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def get_bao_cao_path(name: str) -> Path:
    direct = BAO_CAO_DIR / name
    if direct.exists():
        return direct
    for path in BAO_CAO_DIR.iterdir():
        if path.name == name:
            return path
    return direct


def create_statistics_workbook() -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "TongHop"
    ws.append(["Hạng mục", "Workflow / Runner", "Trigger", "Công cụ", "Quy mô", "Đầu ra chính", "Ghi chú"])
    for row in WORKFLOWS:
        ws.append([row.area, row.workflow, row.trigger, row.tools, row.scope, row.output, row.note])
    style_sheet(ws)

    ws_flows = wb.create_sheet("UI_Flows")
    ws_flows.append(["Mã", "Nhóm", "Test class", "Test method", "Mục tiêu", "Browser"])
    for flow in FLOWS:
        ws_flows.append([flow.flow_id, flow.category, flow.test_class, flow.test_method, flow.objective, flow.browsers])
    style_sheet(ws_flows)

    ws_artifacts = wb.create_sheet("Artifacts")
    ws_artifacts.append(["Nhóm", "Đường dẫn", "Ý nghĩa"])
    for row in ARTIFACT_ROWS:
        ws_artifacts.append(list(row))
    style_sheet(ws_artifacts)

    output = BAO_CAO_DIR / "Bảng thống kê Chương 6 Automation.xlsx"
    wb.save(output)
    return output


def create_testcase_workbook() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Automation TC"
    ws.append(["AT ID", "Nhóm", "Workflow", "Script / Test", "Mục tiêu", "Kết quả mong đợi"])

    rows = [
        ("AT-01", "Backend regression", "backend-regression.yml", "RegressionStableSuite", "Chạy lại suite ổn định trên backend", "Suite chạy thành công và sinh surefire report."),
        ("AT-02", "Performance smoke", "performance-smoke.yml", "BookingPerformanceSmokeTest", "Đo nhanh createBooking và kịch bản đồng thời", "Sinh báo cáo hiệu năng và test pass trong ngưỡng đã đặt."),
        ("AT-03", "Selenium smoke", "selenium-e2e.yml", "FrontendSeleniumSmokeTest", "Chạy UI smoke test trên Chrome", "Workflow pass và upload screenshot, log."),
        ("AT-04", "Selenium smoke", "selenium-e2e.yml", "FrontendSeleniumSmokeTest", "Chạy UI smoke test trên Firefox", "Workflow pass và upload screenshot, log."),
        ("AT-05", "UI flow", "selenium-e2e.yml", "homePageShouldRenderFeaturedTrips", "Trang chủ hiển thị danh sách chuyến", "Tìm thấy trip card và ảnh checkpoint được lưu."),
        ("AT-06", "UI flow", "selenium-e2e.yml", "cashBookingFlowShouldShowSuccessNotification", "Đặt vé CASH thành công", "Hiện thông báo thành công và lưu ảnh booking."),
        ("AT-07", "UI flow", "selenium-e2e.yml", "momoBookingFlowShouldReachPaymentSuccessPage", "Luồng MOMO đi tới trang thành công", "Đi tới Payment Success hoặc My Bookings theo mock flow."),
        ("AT-08", "UI flow", "selenium-e2e.yml", "myBookingsPageShouldRenderExistingBookingsForLoggedInUser", "Hiển thị lịch sử vé", "Trang My Bookings tải được booking card."),
        ("AT-09", "Known issue evidence", "selenium-known-issues.yml", "tc02_sameSeatMomo_shouldRejectOneBooking", "Hai user cùng đặt 1 ghế MOMO", "Nếu cả hai cùng thành công thì testcase fail và vẫn có evidence."),
        ("AT-10", "Known issue evidence", "selenium-known-issues.yml", "tc03_overlapSeatCash_shouldRejectOneBooking", "Hai user đặt ghế giao nhau bằng CASH", "Nếu cả hai cùng thành công thì testcase fail và vẫn có evidence."),
    ]

    for row in rows:
        ws.append(list(row))
    style_sheet(ws)

    output = BAO_CAO_DIR / "TestCase_Chuong6_Automation.xlsx"
    wb.save(output)
    return output


def update_testreport_workbook() -> Path:
    report_path = get_bao_cao_path("TestReport.xlsx")
    workbook = load_workbook(report_path)

    for sheet_name in ["Chuong6_Automation", "Chuong6_Workflows", "Chuong6_Artifacts"]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    ws = workbook.create_sheet("Chuong6_Automation")
    ws.append(["Hạng mục", "Workflow", "Trigger", "Quy mô", "Ghi chú"])
    for row in WORKFLOWS:
        ws.append([row.area, row.workflow, row.trigger, row.scope, row.note])
    style_sheet(ws)

    ws_flows = workbook.create_sheet("Chuong6_Workflows")
    ws_flows.append(["Mã", "Nhóm", "Test class", "Test method", "Mục tiêu"])
    for flow in FLOWS:
        ws_flows.append([flow.flow_id, flow.category, flow.test_class, flow.test_method, flow.objective])
    style_sheet(ws_flows)

    ws_artifacts = workbook.create_sheet("Chuong6_Artifacts")
    ws_artifacts.append(["Nhóm", "Đường dẫn", "Ý nghĩa"])
    for row in ARTIFACT_ROWS:
        ws_artifacts.append(list(row))
    style_sheet(ws_artifacts)

    try:
        workbook.save(report_path)
        return report_path
    except PermissionError:
        fallback = BAO_CAO_DIR / "TestReport_Chuong6_Automation.xlsx"
        workbook.save(fallback)
        return fallback


def create_chapter_docx() -> Path:
    doc = Document()
    for text, style in CHAPTER6_TEXT:
        para = doc.add_paragraph(text)
        para.style = style
        if style == "Heading 1":
            para.alignment = WD_ALIGN_PARAGRAPH.CENTER
            for run in para.runs:
                run.font.size = Pt(16)
                run.bold = True

    output = BAO_CAO_DIR / "Chuong6_KiemThuTuDong_Automation.docx"
    doc.save(output)
    return output


def create_chapter_markdown() -> Path:
    lines = []
    for text, style in CHAPTER6_TEXT:
        if style == "Heading 1":
            lines.append(f"# {text}")
        elif style == "Heading 2":
            lines.append(f"\n## {text}")
        else:
            lines.append(f"\n{text}")
    output = BAO_CAO_DIR / "Chuong6_KiemThuTuDong_Automation.md"
    output.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")
    return output


def update_markdown_inventory(files: list[Path]) -> Path:
    path = BAO_CAO_DIR / "du-lieu-bam-sat-du-an-cho-bao-cao-word.md"
    marker = "\n## Hiện vật Chương 6 Automation\n"
    text = path.read_text(encoding="utf-8")
    block = marker + "\n".join(f"- `{file.name}`" for file in files) + "\n"
    if marker in text:
        text = text.split(marker)[0].rstrip() + "\n" + block
    else:
        text += "\n" + block
    path.write_text(text, encoding="utf-8")
    return path


def main() -> None:
    outputs = [
        create_statistics_workbook(),
        create_testcase_workbook(),
        update_testreport_workbook(),
        create_chapter_docx(),
        create_chapter_markdown(),
    ]
    inventory = update_markdown_inventory(outputs)
    outputs.append(inventory)

    print("Created Chapter 6 supporting files:")
    for output in outputs:
        print(output)


if __name__ == "__main__":
    main()
